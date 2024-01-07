import http
import json
from collections import namedtuple
from datetime import timedelta

import openpyxl
from flask_restx import Namespace, Resource, fields
from sqlalchemy import and_
from werkzeug.utils import secure_filename
from flask_jwt_extended import jwt_required, get_jwt_identity, get_current_user
from ..models.auth_models import *
from ..models.tenders_models import *
from http import HTTPStatus
from ..utils.db import db
from flask import request, send_from_directory
from decouple import config
import os, shutil
from ..verifications.tender_financial_verification import *
from ..verifications.resources_upload_functions import *


import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from pathlib import Path
from email import encoders



tenders_namespace = Namespace('tenders', description="Namespace for tenders")


@tenders_namespace.route('/createTender')
class createTender(Resource):

    @jwt_required()
    def post(self):
        """
            create a new tender
        """
        user_Id = get_jwt_identity()
        data = request.get_json()

        tender = Tenders.query.filter(Tenders.TenderNumber == data['TenderNumber']).first()
        if tender:
            return 'Tender number already exist.', HTTPStatus.CONFLICT

        tenders = Tenders(**data)
        tenders.save()

        activityList = ActivityList.query.order_by(ActivityList.OrderNumber).all()
        for activity in activityList:

            Status = False
            ten_UserListId = None
            # if activity.OrderNumber == 1:
            #     user = UserList.query.filter(UserList.dbo_AspNetUsersId==user_Id).first()
            #     if user:
            #         ten_UserListId = user.Id
            #         Status = True

            tenderActivityRegister = TenderActivityRegister(
                ten_TenderId=tenders.Id,
                ten_ActivityListId=activity.OrderNumber,
                ten_UserListId=ten_UserListId,
                Deadline=None,
                Status=Status,
                CreatedBy=user_Id,
                CreatedOn=datetime.now(),
                UpdatedBy=user_Id,
                UpdatedOn=datetime.now(),
            )

            tenderActivityRegister.save()

        tenderLogs = TenderLogs(
            ten_Tender_Id=tenders.Id,
            FieldName='New Tender Created',
            UpdatedBy_Id=user_Id
        )
        tenderLogs.save()

        return 'Tender Added Successfully.', HTTPStatus.CREATED


@tenders_namespace.route('/updateTender')
class updateTender(Resource):

    @jwt_required()
    def post(self):
        """
            update a tender
        """
        user_Id = get_jwt_identity()
        data = request.get_json()

        Id = data['Id']
        del data['Id']

        tenderSubmissionDates = data['TenderSubmissionDates']
        del data['TenderSubmissionDates']

        lessonsLearned = data['LessonsLearned']
        del data['LessonsLearned']

        planningData = data['PlanningData']
        del data['PlanningData']

        # tenders = Tenders.query.get(Id)
        tenders = Tenders.query.filter_by(Id=Id).first()
        for attribute, value in data.items():

            old_value = getattr(tenders, attribute)
            updated_value = value

            # save the logs for updated field
            if old_value != updated_value:
                tenderLogs = TenderLogs(
                    ten_Tender_Id=Id,
                    FieldName=attribute,
                    ValueBefore=old_value,
                    ValueAfter=updated_value,
                    UpdatedBy_Id=user_Id,
                )
                tenderLogs.save()

            # updating the attribute
            setattr(tenders, attribute, value)

        updated_tick_status_1(tenders, Id)
        updated_tick_status_2(tenders, Id)

        LessonsLearned.query.filter_by(ten_Tender_Id=Id).delete()
        for lesson in lessonsLearned:
            del lesson['Id']
            lessonslearned = LessonsLearned(**lesson)
            lessonslearned.save()

        TenderSubmissionDates.query.filter_by(ten_Tender_Id=Id).delete()
        for tenderSubmission in tenderSubmissionDates:
            # del tenderSubmission['Id']
            tenderSubmissionDate = TenderSubmissionDates(**tenderSubmission)
            tenderSubmissionDate.save()

        db.session.commit()

        if_plannning = False

        Planning.query.filter_by(ten_Tender_Id=Id).delete()
        for planning in planningData:
            del planning['Id']
            planning['DateStart'] = None if planning['DateStart'] is '' else planning['DateStart']
            planning['DateFinish'] = None if planning['DateFinish'] is '' else planning['DateFinish']
            planningData = Planning(**planning)
            planningData.save()

            if_plannning = True

        if if_plannning:
            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == Id,
                TenderActivityRegister.ten_ActivityListId == 5,
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = True

        db.session.commit()

        # updated_object = Tenders.query.filter_by(Id=Id).first()
        # updated_object = updated_object.serialize
        #
        # lessonsLearned = LessonsLearned.query.filter_by(ten_Tender_Id=Id).all()
        # updated_object['LessonsLearned'] = [x.serialize for x in lessonsLearned]

        return 'Tender Updated Successfully.', HTTPStatus.OK


def updated_tick_status_1(tenders, Id):
    tenderActivityRegister = TenderActivityRegister.query.filter(
        TenderActivityRegister.ten_TenderId == Id,
        TenderActivityRegister.ten_ActivityListId == 1,
    ).first()

    if tenders.ProjectName == None or tenders.ProjectName == '':
        return False

    if tenders.TenderNumber == None or tenders.TenderNumber == '':
        return False

    if tenders.Revision == None or tenders.Revision == '':
        return False

    if tenders.Country_Id == None or tenders.Country_Id == '':
        return False

    if tenders.Employer_Id == None or tenders.Employer_Id == '':
        return False

    if tenders.Submission == None or tenders.Submission == '':
        return False

    if tenders.Area == None or tenders.Area == '':
        return False

    if tenders.ScopeBriefFile == None or tenders.ScopeBriefFile == '':
        return False

    if tenders.ScopeNarrative == None or tenders.ScopeNarrative == '':
        return False

    if tenderActivityRegister:
        tenderActivityRegister.Status = True

    return True


def updated_tick_status_2(tenders, Id):
    tenderActivityRegister = TenderActivityRegister.query.filter(
        TenderActivityRegister.ten_TenderId == Id,
        TenderActivityRegister.ten_ActivityListId == 2,
    ).first()

    if tenders.Industry_Id == None or tenders.Industry_Id == '':
        return False

    if tenders.NatureOfWork_Id == None or tenders.NatureOfWork_Id == '':
        return False

    if tenders.TenderDocumentsReceived == None or tenders.TenderDocumentsReceived == '':
        return False

    if tenders.Probability == None or tenders.Probability == '':
        return False

    if tenders.TenderStatus_Id == None or tenders.TenderStatus_Id == '':
        return False

    if tenders.DredgingProdEstimator_Id == None or tenders.DredgingProdEstimator_Id == '':
        return False

    if tenders.PrimaryCostEstimator_Id == None or tenders.PrimaryCostEstimator_Id == '':
        return False

    if tenders.TenderCoordinator_Id == None or tenders.TenderCoordinator_Id == '':
        return False

    if tenders.SubmissionDate == None or tenders.SubmissionDate == '':
        return False

    if tenders.AwardDate == None or tenders.AwardDate == '':
        return False

    if tenderActivityRegister:
        tenderActivityRegister.Status = True

    return True


@tenders_namespace.route('/getTendersList')
class getTendersList(Resource):

    @jwt_required()
    def get(self):
        """
            get all tenders
        """
        tenders = Tenders.query.order_by(Tenders.Id.desc()).all()

        natureOfWorks = NatureOfWorks.query.all()
        natureOfWorks = {x.Id: x.NatureOfWorkName for x in natureOfWorks}

        tenderStatus = TenderStatus.query.all()
        tenderStatus = {x.Id: {'Name': x.TenderStatusName, 'Class': x.ClassName} for x in tenderStatus}

        industry = Industry.query.all()
        industry = {x.Id: x.IndustryName for x in industry}

        country = Country.query.all()
        country = {x.Id: x.CountryName for x in country}

        response_data = []
        for x in tenders:
            x = x.serialize

            try:
                x['NatureOfWork_Id'] = natureOfWorks[x['NatureOfWork_Id']]
            except:
                x['NatureOfWork_Id'] = ''

            try:
                x['TenderStatus_Id'] = tenderStatus[x['TenderStatus_Id']]
            except:
                x['TenderStatus_Id'] = {'Name': '', 'Class': ''}

            try:
                x['Industry_Id'] = industry[x['Industry_Id']]
            except:
                x['Industry_Id'] = ''

            try:
                x['Country_Id'] = country[x['Country_Id']]
            except:
                x['Country_Id'] = ''

            response_data.append(x)

        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getTenderById')
class getTenderById(Resource):

    @jwt_required()
    def get(self):
        """
            get tender by Id
        """
        tender_id = request.args.get('tender_id')
        tender = Tenders.query.filter_by(Id=tender_id).first()

        if tender:
            response_data = tender.serialize
            response_data['TenderSubmissionDates'] = [x.serialize for x in tender.tenderSubmissionDates]
            response_data['LessonsLearned'] = [x.serialize for x in tender.lessonsLearned]
            response_data['TenderFinancialData'] = get_tender_financial_data(tender.tenderFinancialData)
            response_data['TenderEquipmentCostDredging'] = get_tender_equipment_cost_data(tender.tenderEquipmentCost, True)
            response_data['TenderEquipmentCostEstimation'] = get_tender_equipment_cost_data(tender.tenderEquipmentCost, False)
            response_data['Planning'] = get_tender_planning_data(tender.planning)
            response_data['RiskRegister'] = [x.serialize for x in tender.riskRegister]
            response_data['TenderActivityRegister'] = [x.serialize for x in tender.tenderActivityRegister]
            response_data['TenderFinancialSCurve'] = s_curve_fin_perc(tender_id)
            response_data['TenderFinancialSCurvePlan'] = s_curve_fin_plan(tender_id)

            return response_data, HTTPStatus.OK

        return 'Not Found', HTTPStatus.NOT_FOUND


def s_curve_fin_perc(tender_id):
    act_obj = EquipmentCostItem.query.all()
    eq_id_list = []
    parent_id = []
    for act_ins in act_obj:
        if act_ins.SubChild!=None:
            if act_ins.Id not in eq_id_list:
                eq_id_list.append(act_ins.Id)
                if act_ins.ParentId not in parent_id:
                    parent_id.append(act_ins.ParentId)

    for act_ins in act_obj:
        if act_ins.Child!=None:
            if act_ins.Id not in eq_id_list and act_ins.Id not in parent_id:
                eq_id_list.append(act_ins.Id)

    equip_obj = EquipmentCostItem.query.all()
    equip_dict = {}
    for eq_ins in equip_obj:
        equip_dict[eq_ins.Id] = eq_ins.Item

    ten_fin_obj = TenderFinancialData.query.filter(
        TenderFinancialData.ten_TenderId ==tender_id,
        TenderFinancialData.ten_EquipmentCostItemId.in_(eq_id_list)
    ).all()

    total_cost = 0
    for ten_ins in ten_fin_obj:
        total_cost+=ten_ins.TotalCost_inclTax

    x = []
    y = []
    if total_cost > 0:
        for ten_ins in ten_fin_obj:
            x.append(equip_dict[ten_ins.ten_EquipmentCostItemId])
            y.append(ten_ins.TotalCost_inclTax/total_cost)

    return {'x': x, 'y': y}


def s_curve_fin_plan(tender_id):
    act_obj = EquipmentCostItem.query.all()
    eq_id_list = []
    parent_id = []
    for act_ins in act_obj:
        if act_ins.SubChild != None:
            if act_ins.Id not in eq_id_list:
                eq_id_list.append(act_ins.Id)
                if act_ins.ParentId not in parent_id:
                    parent_id.append(act_ins.ParentId)

    for act_ins in act_obj:
        if act_ins.Child != None:
            if act_ins.Id not in eq_id_list and act_ins.Id not in parent_id:
                eq_id_list.append(act_ins.Id)

    equip_obj = EquipmentCostItem.query.all()
    equip_dict = {}
    for eq_ins in equip_obj:
        equip_dict[eq_ins.Id] = eq_ins.Item

    ten_fin_obj = TenderFinancialData.query.filter(
        TenderFinancialData.ten_TenderId == tender_id,
        TenderFinancialData.TotalCost > 0,
        TenderFinancialData.ten_EquipmentCostItemId.in_(eq_id_list)
    ).all()

    total_cost = 0
    for ten_ins in ten_fin_obj:
        total_cost += ten_ins.TotalCost_inclTax

    x = []
    y = []
    x_id = []
    if total_cost > 0:
        for ten_ins in ten_fin_obj:
            x.append(equip_dict[ten_ins.ten_EquipmentCostItemId])
            y.append(ten_ins.TotalCost_inclTax / total_cost)
            x_id.append(ten_ins.ten_EquipmentCostItemId)

    plan_obj = Planning.query.filter(
        Planning.ten_Tender_Id == tender_id,
        Planning.ten_EquipmentCostItemId.in_(x_id)
    ).all()

    plan_dict = {}

    for plan_ins in plan_obj:
        delta = timedelta(days=1)
        start = plan_ins.DateStart
        end = plan_ins.DateFinish

        if start is not None and end is not None:
            # store the dates between two dates in a list
            dates_l = []
            while start <= end:
                dates_l.append(start)
                start += delta

            act_fract = y[x_id.index(plan_ins.ten_EquipmentCostItemId)] / len(dates_l)
            act_name = x[x_id.index(plan_ins.ten_EquipmentCostItemId)]
            for d in dates_l:
                if d.strftime("%Y-%m-%d") not in plan_dict:
                    plan_dict[d.strftime("%Y-%m-%d")] = {name: 0 for name in x}
                plan_dict[d.strftime("%Y-%m-%d")][act_name] += act_fract

    return plan_dict


@tenders_namespace.route('/getTenderNextId')
class getTenderNextId(Resource):

    @jwt_required()
    def get(self):
        """
            get tender by Id
        """
        tender_id = request.args.get('tender_id')
        try:
            if isinstance(tender_id,str):
                next_id = int(tender_id) +1
            else:
                next_id = None
        except:
            next_id = None

        return next_id


def get_tender_financial_data(tenderFinancialData):

    arranged_tenderFinancialData = []
    ten_EquipmentCostItemIds = [x.ten_EquipmentCostItemId for x in tenderFinancialData if x.ten_EquipmentCostItemId != 0 and x.ten_EquipmentCostItemId != None]

    equipmentCostItem = EquipmentCostItem.query.filter(
        EquipmentCostItem.Id.in_(ten_EquipmentCostItemIds)
    ).order_by(
        EquipmentCostItem.Id
    ).all()

    equipmentCostItem = {x.Id: x.serialize for x in equipmentCostItem}

    rows_to_sum = []

    for financialData in tenderFinancialData:

        updated_row = financialData.serialize

        if financialData.ten_EquipmentCostItemId in equipmentCostItem:
            updated_row['AED'] = equipmentCostItem[financialData.ten_EquipmentCostItemId]['ItemReference']
            updated_row['Item'] = equipmentCostItem[financialData.ten_EquipmentCostItemId]['Item']
            updated_row['IsParent'] = equipmentCostItem[financialData.ten_EquipmentCostItemId]['ParentId']

        arranged_tenderFinancialData.append(updated_row)

        if updated_row['AED'] in ['A', 'B', 'C', 'D', 'E']:
            rows_to_sum.append(updated_row)


    arranged_tenderFinancialData = sorted(arranged_tenderFinancialData, key=lambda x: x['AED'])

    if len(rows_to_sum) > 0:
        total = {key: sum(e[key] for e in rows_to_sum if isinstance(e[key], int) or isinstance(e[key], float)) for key in rows_to_sum[0].keys()}
        total['AED'] = ''
        total['Item'] = 'Total'
        total['IsParent'] = False
        arranged_tenderFinancialData.append(total)

    return arranged_tenderFinancialData


def get_tender_equipment_cost_data(tenderEquipmentCost, IsDredging):

    arranged_tenderEquipmentCost = []
    ten_EquipmentItemIds = [x.ten_EquipmentItemId for x in tenderEquipmentCost if
                                x.ten_EquipmentItemId != 0 and x.ten_EquipmentItemId != None and x.IsDredging == IsDredging]

    equipmentItem = EquipmentItem.query.filter(
        EquipmentItem.Id.in_(ten_EquipmentItemIds)
    ).order_by(
        EquipmentItem.Id
    ).all()

    equipmentItem = {str(x.Id): x.serialize for x in equipmentItem}

    for equipmentCost in tenderEquipmentCost:

        if equipmentCost.IsDredging == IsDredging:

            updated_row = equipmentCost.serialize

            del updated_row['CreatedOn']

            if equipmentCost.ten_EquipmentItemId in equipmentItem:
                updated_row['Item'] = equipmentItem[equipmentCost.ten_EquipmentItemId]['Item']
                updated_row['ParentId'] = equipmentItem[equipmentCost.ten_EquipmentItemId]['ParentId']

            arranged_tenderEquipmentCost.append(updated_row)

    arranged_tenderEquipmentCost = sorted(arranged_tenderEquipmentCost, key=lambda x: x['ten_EquipmentItemId'])

    if len(arranged_tenderEquipmentCost) > 0:
        total = {key: sum(e[key] for e in arranged_tenderEquipmentCost if isinstance(e[key], int) or isinstance(e[key], float)) for key
                 in arranged_tenderEquipmentCost[0].keys()}
        total['ten_EquipmentItemId'] = ''
        total['NoOfEquipment'] = ''
        total['ten_Tender_Id'] = ''
        total['Item'] = 'Total'
        arranged_tenderEquipmentCost.append(total)

    return arranged_tenderEquipmentCost


def get_tender_planning_data(tenderPlanningData):

    planning_data = {x.ten_EquipmentCostItemId: x.serialize for x in tenderPlanningData}

    equipmentCostItem = EquipmentCostItem.query.order_by(
        EquipmentCostItem.Id
    ).all()

    response_data = []
    for equipmentCost in equipmentCostItem:

        response_data.append(
            {
                'Id': planning_data[equipmentCost.Id]['Id'] if equipmentCost.Id in planning_data else 0,
                'ten_EquipmentCostItemId': equipmentCost.Id,
                'EquipmentCostItemName': equipmentCost.Item,
                'EquipmentCostItemReference': equipmentCost.ItemReference,
                'DateStart': planning_data[equipmentCost.Id]['DateStart'] if equipmentCost.Id in planning_data else '',
                'DateFinish': planning_data[equipmentCost.Id]['DateFinish'] if equipmentCost.Id in planning_data else '',
            }
        )

    return response_data


@tenders_namespace.route('/uploadTenderFinancialData')
class uploadTenderFinancialData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender financial data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderFinancialExcelFiles'))

        if 'financial_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['financial_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_PATH, 'tenderFinancialExcelFiles/{}'.format(filename)))

        validation = vailadate_financial_excel_file(os.path.join(UPLOAD_PATH, 'tenderFinancialExcelFiles/{}'.format(filename)))

        if validation.jsonerror_counter == 0:
            load_financial_data_json(validation.json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 3,
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = True
                db.session.commit()

        # adding work load columns values

        equipmentCostItem = EquipmentCostItem.query.filter(
            EquipmentCostItem.ItemReference.in_(
                ['C.1', 'C.2', 'C.3', 'C.4', 'C.5', 'C.6', 'C.7', 'C.8', 'C.9', 'C.10', 'C.11', 'C.12', 'C.13'])
        ).all()

        _dredging = [x.Id for x in equipmentCostItem if x.ItemReference in ['C.1', 'C.2']]
        _marineCivil = [x.Id for x in equipmentCostItem if
                        x.ItemReference in ['C.3', 'C.4', 'C.5', 'C.6', 'C.7', 'C.8', 'C.10', 'C.11', 'C.12', 'C.13']]
        _gi = [x.Id for x in equipmentCostItem if x.ItemReference in ['C.9']]

        Dredging = 0
        MarineCivil = 0
        GI = 0

        tenderFinancialData = TenderFinancialData.query.filter(
            TenderFinancialData.ten_TenderId == tender_id,
            TenderFinancialData.ten_EquipmentCostItemId.in_([x.Id for x in equipmentCostItem])
        ).all()

        for tenderFinancial in tenderFinancialData:
            if tenderFinancial.ten_EquipmentCostItemId in _dredging:
                Dredging += tenderFinancial.EstimatedCost if tenderFinancial.EstimatedCost is not None else 0

            if tenderFinancial.ten_EquipmentCostItemId in _marineCivil:
                MarineCivil += tenderFinancial.EstimatedCost if tenderFinancial.EstimatedCost is not None else 0

            if tenderFinancial.ten_EquipmentCostItemId in _gi:
                GI += tenderFinancial.EstimatedCost if tenderFinancial.EstimatedCost is not None else 0

        Tenders.query.filter(
            Tenders.Id == tender_id
        ).update(
            {
                'CivilMarineEPC': MarineCivil,
                'GIPillingWorks': GI,
                'Dredging': Dredging,
            }
        )

        db.session.commit()


        return {
            'json_dict': validation.json_error,
            'jsonerror_counter': validation.jsonerror_counter
        }, HTTPStatus.OK


def vailadate_financial_excel_file(file_path):
    dataformtender_obj = new_LoadTenderDataForm(file_path)
    dataformtender_obj.financial_data()
    return dataformtender_obj


def delete_files_from_folder(folder):
    for filename in os.listdir(folder):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print('Failed to delete %s. Reason: %s' % (file_path, e))


@tenders_namespace.route('/deleteTender')
class deleteTender(Resource):

    @jwt_required()
    def get(self):
        """
            delete tender
        """

        tender_id = request.args.get('tender_id')

        LessonsLearned.query.filter_by(ten_Tender_Id=tender_id).delete()
        TenderFinancialData.query.filter_by(ten_TenderId=tender_id).delete()
        TenderEquipmentCost.query.filter_by(ten_Tender_Id=tender_id).delete()
        Planning.query.filter_by(ten_Tender_Id=tender_id).delete()
        TenderLogs.query.filter_by(ten_Tender_Id=tender_id).delete()
        RiskRegister.query.filter_by(ten_TenderId=tender_id).delete()
        TenderActivityRegister.query.filter_by(ten_TenderId=tender_id).delete()
        Tenders.query.filter_by(Id=tender_id).delete()

        db.session.commit()

        return 'Tender deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/uploadTenderEquipmentDredgingData')
class uploadTenderEquipmentDredgingData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender financial data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderEquipmentDredgingFiles'))

        if 'equipment_dredging_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['equipment_dredging_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_PATH, 'tenderEquipmentDredgingFiles/{}'.format(filename)))

        validation = vailadate_equipment_dredging_excel_file(os.path.join(UPLOAD_PATH, 'tenderEquipmentDredgingFiles/{}'.format(filename)))

        if validation.jsonerror_dform_counter == 0:
            load_dredging_eq_data_json(validation.json_dataform_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 4,
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = True
                db.session.commit()

        return {
            'json_dict': validation.json_dataform_error,
            'jsonerror_counter': validation.jsonerror_dform_counter
        }, HTTPStatus.OK


def vailadate_equipment_dredging_excel_file(file_path):
    dataformtender_obj = LoadEquipmentDredging(file_path)
    dataformtender_obj.dredging_equip_data()
    return dataformtender_obj


@tenders_namespace.route('/uploadTenderEquipmentEstimationData')
class uploadTenderEquipmentEstimationData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender financial data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderEquipmentEstimationFiles'))

        if 'equipment_estimation_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['equipment_estimation_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_PATH, 'tenderEquipmentEstimationFiles/{}'.format(filename)))

        validation = vailadate_equipment_estimation_excel_file(os.path.join(UPLOAD_PATH, 'tenderEquipmentEstimationFiles/{}'.format(filename)))

        if validation.json_eq_counter == 0:
            load_estimator_eq_data_json(validation.json_eq_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 4,
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = True
                db.session.commit()

        return {
            'json_dict': validation.json_eq_error,
            'jsonerror_counter': validation.json_eq_counter
        }, HTTPStatus.OK


def vailadate_equipment_estimation_excel_file(file_path):
    dataformtender_obj = LoadEquipmentCostEst(file_path)
    dataformtender_obj.estimation_equip_data()
    return dataformtender_obj


@tenders_namespace.route('/uploadTenderPlanningData')
class uploadTenderPlanningData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender planning data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderPlanningExcelFiles'))

        if 'planning_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['planning_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_PATH, 'tenderPlanningExcelFiles/{}'.format(filename)))

        validation = vailadate_planning_excel_file(os.path.join(UPLOAD_PATH, 'tenderPlanningExcelFiles/{}'.format(filename)), tender_id)

        if validation.json_pl_counter == 0:
            load_planning_data_json(validation.json_pl_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 5,
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = True
                db.session.commit()

        return {
            'json_dict': validation.json_pl_error,
            'jsonerror_counter': validation.json_pl_counter
        }, HTTPStatus.OK


def vailadate_planning_excel_file(file_path, tender_id):
    dataformtender_obj = LoadTenderPlanning(file_path)
    dataformtender_obj.planning_data(tender_id)
    return dataformtender_obj


@tenders_namespace.route('/updateRiskRegister')
class updateRiskRegister(Resource):

    @jwt_required()
    def post(self):
        """
            update Risk Register
        """
        user_Id = get_jwt_identity()
        data = request.get_json()

        Id = data['Id']
        del data['Id']

        data['IsRisk'] = True if data['IsRisk'] == '1' else False
        data['IsFixed'] = True if data['IsFixed'] == '1' else False

        if Id == '0':

            riskRegister = RiskRegister(**data)
            riskRegister.CreatedOn = datetime.now()
            riskRegister.UpdatedOn = datetime.now()
            riskRegister.CreatedBy = user_Id
            riskRegister.UpdatedBy = user_Id
            riskRegister.save()

            db.session.commit()
            return 'Risk Register Inserted Successfully.', HTTPStatus.OK
        else:
            riskRegister = RiskRegister.query.filter(RiskRegister.Id==Id).first()
            if riskRegister:
                riskRegister.UpdatedOn = datetime.now()
                riskRegister.UpdatedBy = user_Id
                riskRegister.IsRisk = data['IsRisk']
                riskRegister.IsFixed = data['IsFixed']
                riskRegister.ten_RiskCategory = data['ten_RiskCategory']
                riskRegister.RiskElement = data['RiskElement']
                riskRegister.EventDetails = data['EventDetails']
                riskRegister.ConsequenceDetails = data['ConsequenceDetails']
                riskRegister.ten_LevelOfRiskId_initial = data['ten_LevelOfRiskId_initial']
                riskRegister.TreatmentMitigation = data['TreatmentMitigation']
                riskRegister.ten_LevelOfRiskId_residual = data['ten_LevelOfRiskId_residual']
                riskRegister.ImplementationStatus = data['ImplementationStatus']
                riskRegister.RiskValue = data['RiskValue']
                riskRegister.Probability = data['Probability']
                riskRegister.RiskAllowance = data['RiskAllowance']

                db.session.commit()
            return 'Risk Register Updated Successfully.', HTTPStatus.OK


@tenders_namespace.route('/getRiskRegisterbyId')
class getRiskRegisterbyId(Resource):

    @jwt_required()
    def get(self):
        """
            get risk register by Id
        """
        risk_id = request.args.get('risk_id')
        riskRegister = RiskRegister.query.filter_by(Id=risk_id).first()

        if riskRegister:
            return riskRegister.serialize, HTTPStatus.OK

        return 'Risk Register Not Found.', HTTPStatus.OK


@tenders_namespace.route('/deleteRiskRegister')
class deleteRiskRegister(Resource):

    @jwt_required()
    def get(self):
        """
            delete risk register
        """
        risk_id = request.args.get('risk_id')
        RiskRegister.query.filter_by(Id=risk_id).delete()
        db.session.commit()

        return 'Risk Register deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/getTenderActivitiesList')
class getTenderActivitiesList(Resource):

    @jwt_required()
    def get(self):
        """
            get tender activities
        """
        tender_id = request.args.get('tender_id')
        tenderActivityRegister = TenderActivityRegister.query.filter(TenderActivityRegister.ten_TenderId==tender_id).all()
        return [x.serialize for x in tenderActivityRegister], HTTPStatus.OK


@tenders_namespace.route('/updateTenderActivityUser')
class updateTenderActivityUser(Resource):

    @jwt_required()
    def post(self):
        """
            update tender activity user
        """
        data = request.get_json()

        Id = data['Id']
        Value = data['Value']

        tenderActivityRegister = TenderActivityRegister.query.filter(TenderActivityRegister.Id==Id).first()
        if tenderActivityRegister:
            tenderActivityRegister.ten_UserListId = Value
            tenderActivityRegister.UpdatedOn = datetime.now()
            tenderActivityRegister.UpdatedBy = get_jwt_identity()
            db.session.commit()

            return {
                'success': 'Activity User Updated Successfully.',
                'last_updated': str(tenderActivityRegister.UpdatedOn),
            }, HTTPStatus.OK

        return 'Not Found', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/updateTenderActivityDeadline')
class updateTenderActivityDeadline(Resource):

    @jwt_required()
    def post(self):
        """
            update tender activity user
        """
        data = request.get_json()

        Id = data['Id']
        Value = data['Value']

        tenderActivityRegister = TenderActivityRegister.query.filter(TenderActivityRegister.Id==Id).first()
        if tenderActivityRegister:
            tenderActivityRegister.Deadline = Value
            tenderActivityRegister.UpdatedOn = datetime.now()
            tenderActivityRegister.UpdatedBy = get_jwt_identity()
            db.session.commit()

            return {
                       'success': 'Activity User Updated Successfully.',
                       'last_updated': str(tenderActivityRegister.UpdatedOn),
                   }, HTTPStatus.OK

        return 'Not Found', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/getActivitiesbyTenderId')
class getActivitiesbyTenderId(Resource):

    @jwt_required()
    def get(self):
        """
            get activities by Tender Id
        """
        tender_id = request.args.get('tender_id')
        tenderActivityRegister = TenderActivityRegister.query.filter_by(ten_TenderId=tender_id).all()
        return [x.serialize for x in tenderActivityRegister], HTTPStatus.OK


@tenders_namespace.route('/updateUserPosition')
class updateUserPosition(Resource):

    @jwt_required()
    def post(self):
        """
            update user position
        """
        data = request.get_json()

        user_id = data['user_id'].split(',')
        position_id = data['position_id']

        userList = UserList.query.filter(UserList.ten_UserPositionListId == position_id).all()

        # remove old positions, if not selected
        for user in userList:
            if str(user.Id) in user_id:
                check = 'keep'
            else:
                user.ten_UserPositionListId = None
                db.session.commit()

        # add positions, if new is selected
        for user in user_id:
            userList = UserList.query.filter(UserList.Id==user).first()
            if userList:
                userList.ten_UserPositionListId = position_id
                db.session.commit()

        return 'User position updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/uploadTenderKMZFile')
class uploadTenderKMZFile(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender KMZ file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'kmz_file' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['kmz_file']

        if file.filename == '':
            return 'File Not Found.', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)
        if file_extension != '.kmz':
            return 'File extension is not valid.', HTTPStatus.FORBIDDEN

        tender = Tenders.query.filter_by(Id=tender_id).first()
        if tender:
            if tender.KMZFile is not None:
                if os.path.exists(os.path.join(UPLOAD_PATH, 'tenderKMZFiles/{}'.format(tender.KMZFile))):
                    os.remove(os.path.join(UPLOAD_PATH, 'tenderKMZFiles/{}'.format(tender.KMZFile)))
        else:
            return 'Tender Not Found.', HTTPStatus.NOT_FOUND

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + '.kmz'
        file.save(os.path.join(UPLOAD_PATH, 'tenderKMZFiles/{}'.format(filename)))
        tender.KMZFile = filename
        db.session.commit()

        return {
            'filename': tender.KMZFile,
            'foldername': 'tenderKMZFiles',
        }, HTTPStatus.OK


@tenders_namespace.route('/getFile')
class getFile(Resource):

    def get(self):
        """
            get file from server
        """
        filename = request.args.get('filename')
        foldername = request.args.get('foldername')

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        uploads = os.path.join(BASE_DIR, 'static\\' + foldername)

        if os.path.exists(os.path.join(uploads, filename)):
            response = send_from_directory(directory=uploads, path=filename, as_attachment=True)
            response.headers['Content-Type'] = 'application/vnd'
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            return response
        else:
            return 'File not found.', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/uploadTenderFile')
class uploadTenderFile(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender pdf file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'pdf_file' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        attribute_name = request.form['attribute_name']
        folder_name = request.form['folder_name']
        file = request.files['pdf_file']
        extensions = request.form['extensions'].replace(' ','').split(',')

        if file.filename == '':
            return 'File Not Found.', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)
        if file_extension not in extensions:
            return 'File extension is not valid.', HTTPStatus.FORBIDDEN

        tender = Tenders.query.filter_by(Id=tender_id).first()
        if tender:
            old_value = getattr(tender, attribute_name)
            if old_value is not None:
                if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format(folder_name, old_value))):
                    os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format(folder_name, old_value)))
        else:
            return 'Tender Not Found.', HTTPStatus.NOT_FOUND

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format(folder_name, filename)))
        setattr(tender, attribute_name, filename)
        db.session.commit()

        return {
            'status': attribute_name + ' uploaded successully.',
            'filename': getattr(tender, attribute_name),
            'foldername': folder_name,
        }, HTTPStatus.OK


@tenders_namespace.route('/addUpdateUserPosition')
class addUpdateUserPosition(Resource):

    @jwt_required()
    def post(self):
        """
            add/update user position
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']

        if Id == '0':
            userPositionList = UserPositionList.query.filter(
                UserPositionList.Name == Name,
            ).first()
            if userPositionList is not None:
                return 'User position exist already.', HTTPStatus.CONFLICT

            obj = UserPositionList(
                Name=Name,
                CreatedOn=datetime.now(),
                CreatedBy=get_jwt_identity(),
                UpdatedOn=datetime.now(),
                UpdatedBy=get_jwt_identity(),
            )
            obj.save()

            return 'User position registered successfully.', HTTPStatus.OK

        userPositionList = UserPositionList.query.filter(UserPositionList.Id==Id).first()
        if userPositionList:
            userPositionList.Name = Name
            userPositionList.UpdatedOn = datetime.now()
            userPositionList.UpdatedBy = get_jwt_identity()
            db.session.commit()

        return 'User position updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteUserPosition')
class deleteUserPosition(Resource):

    @jwt_required()
    def get(self):
        """
            delete user position
        """
        position_id = request.args.get('position_id')
        userList = UserList.query.filter_by(ten_UserPositionListId=position_id).all()
        for user in userList:
            user.ten_UserPositionListId = None
            user.UpdatedOn = datetime.now()
            user.UpdatedBy = get_jwt_identity()
        UserPositionList.query.filter_by(Id=position_id).delete()
        db.session.commit()
        return 'User position deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/uploadTenderRiskRegisterData')
class uploadTenderRiskRegisterData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender risk register data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderRiskRegisterExcelFiles'))

        if 'risk_register_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['risk_register_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        file.save(os.path.join(UPLOAD_PATH, 'tenderRiskRegisterExcelFiles/{}'.format(filename)))

        data = upload_risk_register_excel(os.path.join(UPLOAD_PATH, 'tenderRiskRegisterExcelFiles/{}'.format(filename)), tender_id, get_jwt_identity())

        if len(data['json_error']) == 0:
            for obj in data['data_to_insert']:
                object_name = RiskRegister(**obj)
                object_name.save()

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 9
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        return {
            'json_dict': data['json_error'],
            'jsonerror_counter': len(data['json_error'])
        }, HTTPStatus.OK


@tenders_namespace.route('/updateResourceManpower')
class updateResourceManpower(Resource):

    @jwt_required()
    def post(self):
        """
            update resource manpower
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        ten_ResourceManpowerLevelId_1 = data['ten_ResourceManpowerLevelId_1']
        ten_ResourceManpowerLevelId_2 = data['ten_ResourceManpowerLevelId_2']
        ten_ResourceManpowerLevelId_3 = data['ten_ResourceManpowerLevelId_3']
        ten_ResourceManpowerLevelId_4 = data['ten_ResourceManpowerLevelId_4']
        ten_ResourceManpowerLevelId_5 = data['ten_ResourceManpowerLevelId_5']
        ten_ResourceDepartmentId = data['ten_ResourceDepartmentId']
        ten_ResourceCodeId = data['ten_ResourceCodeId']

        if Id == '0':
            resourceManpowerMain = ResourceManpower.query.filter(ResourceManpower.Name == Name).first()
            if resourceManpowerMain is not None:
                return 'Manpower exist already.', HTTPStatus.CONFLICT

            obj = ResourceManpower(
                Name=Name,
                ten_ResourceManpowerLevelId_1=None if ten_ResourceManpowerLevelId_1 is '' else ten_ResourceManpowerLevelId_1,
                ten_ResourceManpowerLevelId_2=None if ten_ResourceManpowerLevelId_2 is '' else ten_ResourceManpowerLevelId_2,
                ten_ResourceManpowerLevelId_3=None if ten_ResourceManpowerLevelId_3 is '' else ten_ResourceManpowerLevelId_3,
                ten_ResourceManpowerLevelId_4=None if ten_ResourceManpowerLevelId_4 is '' else ten_ResourceManpowerLevelId_4,
                ten_ResourceManpowerLevelId_5=None if ten_ResourceManpowerLevelId_5 is '' else ten_ResourceManpowerLevelId_5,
                ten_ResourceDepartmentId=ten_ResourceDepartmentId,
                ten_ResourceCodeId=ten_ResourceCodeId
            )
            obj.save()

            return 'Manpower registered successfully.', HTTPStatus.OK

        resourceManpowerMain = ResourceManpower.query.filter(ResourceManpower.Id==Id).first()
        if resourceManpowerMain:
            resourceManpowerMain.Name = Name
            resourceManpowerMain.ten_ResourceManpowerLevelId_1 = None if ten_ResourceManpowerLevelId_1 is '' else ten_ResourceManpowerLevelId_1
            resourceManpowerMain.ten_ResourceManpowerLevelId_2 = None if ten_ResourceManpowerLevelId_2 is '' else ten_ResourceManpowerLevelId_2
            resourceManpowerMain.ten_ResourceManpowerLevelId_3 = None if ten_ResourceManpowerLevelId_3 is '' else ten_ResourceManpowerLevelId_3
            resourceManpowerMain.ten_ResourceManpowerLevelId_4 = None if ten_ResourceManpowerLevelId_4 is '' else ten_ResourceManpowerLevelId_4
            resourceManpowerMain.ten_ResourceManpowerLevelId_5 = None if ten_ResourceManpowerLevelId_5 is '' else ten_ResourceManpowerLevelId_5
            resourceManpowerMain.ten_ResourceDepartmentId = ten_ResourceDepartmentId
            resourceManpowerMain.ten_ResourceCodeId = ten_ResourceCodeId
            db.session.commit()

        return 'Manpower updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceManpower')
class deleteResourceManpower(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource manpower
        """
        id = request.args.get('id')
        ResourceManpower.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource manpower deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/getResourceManpower')
class getResourceManpower(Resource):

    @jwt_required()
    def get(self):
        """
            get resource manpower
        """
        resourceManpower = ResourceManpower.query.all()
        resourceManpowerLevel = ResourceManpowerLevel.query.all()
        resourceManpowerLevelDist = {}
        for manpowerLevel in resourceManpowerLevel:
            if manpowerLevel.LevelNumber not in resourceManpowerLevelDist:
                resourceManpowerLevelDist[manpowerLevel.LevelNumber] = {}

            if manpowerLevel.Id not in resourceManpowerLevelDist[manpowerLevel.LevelNumber]:
                resourceManpowerLevelDist[manpowerLevel.LevelNumber][manpowerLevel.Id] = manpowerLevel.Name

        resourceDepartment = ResourceDepartment.query.all()
        resourceDepartment = {x.Id: x.Name for x in resourceDepartment}

        resourceCode = ResourceCode.query.all()
        resourceCode = {x.Id: x.Name for x in resourceCode}

        response_data = []
        for manpower in resourceManpower:
            obj = manpower.serialize

            try:
                obj['ten_ResourceDepartmentIdName'] = resourceDepartment[manpower.ten_ResourceDepartmentId]
            except:
                obj['ten_ResourceDepartmentIdName'] = ''

            try:
                obj['ten_ResourceCodeIdName'] = resourceCode[manpower.ten_ResourceCodeId]
            except:
                obj['ten_ResourceCodeIdName'] = ''

            try:
                obj['ten_ResourceManpowerLevelId_1Name'] = resourceManpowerLevelDist[1][manpower.ten_ResourceManpowerLevelId_1]
            except:
                obj['ten_ResourceManpowerLevelId_1Name'] = ''

            try:
                obj['ten_ResourceManpowerLevelId_2Name'] = resourceManpowerLevelDist[2][manpower.ten_ResourceManpowerLevelId_2]
            except:
                obj['ten_ResourceManpowerLevelId_2Name'] = ''

            try:
                obj['ten_ResourceManpowerLevelId_3Name'] = resourceManpowerLevelDist[3][manpower.ten_ResourceManpowerLevelId_3]
            except:
                obj['ten_ResourceManpowerLevelId_3Name'] = ''

            try:
                obj['ten_ResourceManpowerLevelId_4Name'] = resourceManpowerLevelDist[4][manpower.ten_ResourceManpowerLevelId_4]
            except:
                obj['ten_ResourceManpowerLevelId_4Name'] = ''

            try:
                obj['ten_ResourceManpowerLevelId_5Name'] = resourceManpowerLevelDist[5][manpower.ten_ResourceManpowerLevelId_5]
            except:
                obj['ten_ResourceManpowerLevelId_5Name'] = ''

            response_data.append(obj)

        return response_data, HTTPStatus.OK


@tenders_namespace.route('/updateResourceEquipment')
class updateResourceEquipment(Resource):

    @jwt_required()
    def post(self):
        """
            update resource equipment
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        ten_ResourceEquipmentLevelId_1 = data['ten_ResourceEquipmentLevelId_1']
        ten_ResourceEquipmentLevelId_2 = data['ten_ResourceEquipmentLevelId_2']
        ten_ResourceEquipmentLevelId_3 = data['ten_ResourceEquipmentLevelId_3']
        ten_ResourceDepartmentId = data['ten_ResourceDepartmentId']
        ten_ResourceCodeId = data['ten_ResourceCodeId']

        if Id == '0':
            resourceEquipmentMain = ResourceEquipment.query.filter(ResourceEquipment.Name == Name).first()
            if resourceEquipmentMain is not None:
                return 'Equipment exist already.', HTTPStatus.CONFLICT

            obj = ResourceEquipment(
                Name=Name,
                ten_ResourceEquipmentLevelId_1=None if ten_ResourceEquipmentLevelId_1 is '' else ten_ResourceEquipmentLevelId_1,
                ten_ResourceEquipmentLevelId_2=None if ten_ResourceEquipmentLevelId_2 is '' else ten_ResourceEquipmentLevelId_2,
                ten_ResourceEquipmentLevelId_3=None if ten_ResourceEquipmentLevelId_3 is '' else ten_ResourceEquipmentLevelId_3,
                ten_ResourceDepartmentId=ten_ResourceDepartmentId,
                ten_ResourceCodeId=ten_ResourceCodeId,
            )
            obj.save()

            return 'Equipment registered successfully.', HTTPStatus.OK

        resourceEquipment = ResourceEquipment.query.filter(ResourceEquipment.Id==Id).first()
        if resourceEquipment:
            resourceEquipment.Name = Name
            resourceEquipment.ten_ResourceEquipmentLevelId_1 = None if ten_ResourceEquipmentLevelId_1 is '' else ten_ResourceEquipmentLevelId_1
            resourceEquipment.ten_ResourceEquipmentLevelId_2 = None if ten_ResourceEquipmentLevelId_2 is '' else ten_ResourceEquipmentLevelId_2
            resourceEquipment.ten_ResourceEquipmentLevelId_3 = None if ten_ResourceEquipmentLevelId_3 is '' else ten_ResourceEquipmentLevelId_3
            resourceEquipment.ten_ResourceDepartmentId = ten_ResourceDepartmentId
            resourceEquipment.ten_ResourceCodeId = ten_ResourceCodeId
            db.session.commit()

        return 'Equipment updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceEquipment')
class deleteResourceEquipment(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource equipment
        """
        id = request.args.get('id')
        ResourceEquipment.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource equipment deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/getResourceEquipment')
class getResourceEquipment(Resource):

    @jwt_required()
    def get(self):
        """
            get resource equipment
        """
        resourceEquipment = ResourceEquipment.query.all()
        resourceEquipmentLevel = ResourceEquipmentLevel.query.all()
        resourceEquipmentLevelDist = {}
        for equipmentLevel in resourceEquipmentLevel:
            if equipmentLevel.LevelNumber not in resourceEquipmentLevelDist:
                resourceEquipmentLevelDist[equipmentLevel.LevelNumber] = {}

            if equipmentLevel.Id not in resourceEquipmentLevelDist[equipmentLevel.LevelNumber]:
                resourceEquipmentLevelDist[equipmentLevel.LevelNumber][equipmentLevel.Id] = equipmentLevel.Name

        resourceDepartment = ResourceDepartment.query.all()
        resourceDepartment = {x.Id: x.Name for x in resourceDepartment}

        resourceCode = ResourceCode.query.all()
        resourceCode = {x.Id: x.Name for x in resourceCode}

        response_data = []

        for equipment in resourceEquipment:
            obj = equipment.serialize

            try:
                obj['ten_ResourceDepartmentIdName'] = resourceDepartment[equipment.ten_ResourceDepartmentId]
            except:
                obj['ten_ResourceDepartmentIdName'] = ''

            try:
                obj['ten_ResourceCodeIdName'] = resourceCode[equipment.ten_ResourceCodeId]
            except:
                obj['ten_ResourceCodeIdName'] = ''

            try:
                obj['ten_ResourceEquipmentLevelId_1Name'] = resourceEquipmentLevelDist[1][equipment.ten_ResourceEquipmentLevelId_1]
            except:
                obj['ten_ResourceEquipmentLevelId_1Name'] = ''

            try:
                obj['ten_ResourceEquipmentLevelId_2Name'] = resourceEquipmentLevelDist[2][equipment.ten_ResourceEquipmentLevelId_2]
            except:
                obj['ten_ResourceEquipmentLevelId_2Name'] = ''

            try:
                obj['ten_ResourceEquipmentLevelId_3Name'] = resourceEquipmentLevelDist[3][equipment.ten_ResourceEquipmentLevelId_3]
            except:
                obj['ten_ResourceEquipmentLevelId_3Name'] = ''

            response_data.append(obj)

        return response_data, HTTPStatus.OK


@tenders_namespace.route('/updateResourceMaterial')
class updateResourceMaterial(Resource):

    @jwt_required()
    def post(self):
        """
            update resource material
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        ten_ResourceMaterialLevelId_1 = data['ten_ResourceMaterialLevelId_1']
        dbo_UnitId = data['dbo_UnitId']

        if Id == '0':
            resourceMaterial = ResourceMaterial.query.filter(ResourceMaterial.Name == Name).first()
            if resourceMaterial is not None:
                return 'Material exist already.', HTTPStatus.CONFLICT

            obj = ResourceMaterial(
                Name=Name,
                ten_ResourceMaterialLevelId_1=None if ten_ResourceMaterialLevelId_1 is '' else ten_ResourceMaterialLevelId_1,
                dbo_UnitId=None if dbo_UnitId is '' else dbo_UnitId
            )
            obj.save()

            return 'Material registered successfully.', HTTPStatus.OK

        resourceMaterial = ResourceMaterial.query.filter(ResourceMaterial.Id==Id).first()
        if resourceMaterial:
            resourceMaterial.Name = Name
            resourceMaterial.ten_ResourceMaterialLevelId_1 = None if ten_ResourceMaterialLevelId_1 is '' else ten_ResourceMaterialLevelId_1
            resourceMaterial.dbo_UnitId = None if dbo_UnitId is '' else dbo_UnitId
            db.session.commit()

        return 'Material updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceMaterial')
class deleteResourceMaterial(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource material
        """
        id = request.args.get('id')
        ResourceMaterial.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource material deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/getResourceMaterial')
class getResourceMaterial(Resource):

    @jwt_required()
    def get(self):
        """
            get resource material
        """
        resourceMaterial = ResourceMaterial.query.all()
        resourceMaterialLevel = ResourceMaterialLevel.query.all()
        resourceMaterialLevelDist = {}
        for materialLevel in resourceMaterialLevel:
            if materialLevel.LevelNumber not in resourceMaterialLevelDist:
                resourceMaterialLevelDist[materialLevel.LevelNumber] = {}

            if materialLevel.Id not in resourceMaterialLevelDist[materialLevel.LevelNumber]:
                resourceMaterialLevelDist[materialLevel.LevelNumber][materialLevel.Id] = materialLevel.Name

        unit = Unit.query.all()
        unit = {x.Id: x.Name for x in unit}

        response_data = []

        for material in resourceMaterial:
            obj = material.serialize

            try:
                obj['ten_ResourceMaterialLevelId_1Name'] = resourceMaterialLevelDist[1][material.ten_ResourceMaterialLevelId_1]
            except:
                obj['ten_ResourceMaterialLevelId_1Name'] = ''

            try:
                obj['dbo_UnitIdName'] = unit[material.dbo_UnitId]
            except:
                obj['dbo_UnitIdName'] = ''

            response_data.append(obj)

        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getResourceManpowerLevel')
class getResourceManpowerLevel(Resource):

    @jwt_required()
    def get(self):
        """
            get resource manpower level
        """
        resourceManpowerLevel = ResourceManpowerLevel.query.all()
        response_data = [x.serialize for x in resourceManpowerLevel]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getResourceEquipmentLevel')
class getResourceEquipmentLevel(Resource):

    @jwt_required()
    def get(self):
        """
            get resource equipment level
        """
        resourceEquipmentLevel = ResourceEquipmentLevel.query.all()
        response_data = [x.serialize for x in resourceEquipmentLevel]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getResourceMaterialLevel')
class getResourceMaterialLevel(Resource):

    @jwt_required()
    def get(self):
        """
            get resource material level
        """
        resourceMaterialLevel = ResourceMaterialLevel.query.all()
        response_data = [x.serialize for x in resourceMaterialLevel]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getResourceDepartment')
class getResourceDepartment(Resource):

    @jwt_required()
    def get(self):
        """
            get resource department
        """
        resourceDepartment = ResourceDepartment.query.all()
        response_data = [x.serialize for x in resourceDepartment]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getResourceCode')
class getResourceCode(Resource):

    @jwt_required()
    def get(self):
        """
            get resource code
        """
        resourceCode = ResourceCode.query.all()
        response_data = [x.serialize for x in resourceCode]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/getUnit')
class getUnit(Resource):

    @jwt_required()
    def get(self):
        """
            get unit
        """
        unit = Unit.query.all()
        response_data = [x.serialize for x in unit]
        return response_data, HTTPStatus.OK


@tenders_namespace.route('/updateResourceManpowerLevel')
class updateResourceManpowerLevel(Resource):

    @jwt_required()
    def post(self):
        """
            update resource manpower level
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        LevelNumber = data['LevelNumber']

        if Id == '0':
            resourceManpowerLevel = ResourceManpowerLevel.query.filter(ResourceManpowerLevel.Name == Name).first()
            if resourceManpowerLevel is not None:
                return 'Manpower level exist already.', HTTPStatus.CONFLICT

            obj = ResourceManpowerLevel(
                Name=Name,
                LevelNumber=LevelNumber
            )
            obj.save()

            return 'Manpower level registered successfully.', HTTPStatus.OK

        resourceManpowerLevel = ResourceManpowerLevel.query.filter(ResourceManpowerLevel.Id==Id).first()
        if resourceManpowerLevel:
            resourceManpowerLevel.Name = Name
            resourceManpowerLevel.LevelNumber = LevelNumber
            db.session.commit()

        return 'Manpower level updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceManpowerLevel')
class deleteResourceManpowerLevel(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource manpower level
        """
        id = request.args.get('id')
        ResourceManpowerLevel.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource manpower level deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/updateResourceEquipmentLevel')
class updateResourceEquipmentLevel(Resource):

    @jwt_required()
    def post(self):
        """
            update resource equipment level
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        LevelNumber = data['LevelNumber']

        if Id == '0':
            resourceEquipmentLevel = ResourceEquipmentLevel.query.filter(ResourceEquipmentLevel.Name == Name).first()
            if resourceEquipmentLevel is not None:
                return 'Equipment level exist already.', HTTPStatus.CONFLICT

            obj = ResourceEquipmentLevel(
                Name=Name,
                LevelNumber=LevelNumber
            )
            obj.save()

            return 'Equipment level registered successfully.', HTTPStatus.OK

        resourceEquipmentLevel = ResourceEquipmentLevel.query.filter(ResourceEquipmentLevel.Id==Id).first()
        if resourceEquipmentLevel:
            resourceEquipmentLevel.Name = Name
            resourceEquipmentLevel.LevelNumber = LevelNumber
            db.session.commit()

        return 'Equipment level updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceEquipmentLevel')
class deleteResourceEquipmentLevel(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource equipment level
        """
        id = request.args.get('id')
        ResourceEquipmentLevel.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource equipment level deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/updateResourceMaterialLevel')
class updateResourceMaterialLevel(Resource):

    @jwt_required()
    def post(self):
        """
            update resource material level
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        LevelNumber = data['LevelNumber']

        if Id == '0':
            resourceMaterialLevel = ResourceMaterialLevel.query.filter(ResourceMaterialLevel.Name == Name).first()
            if resourceMaterialLevel is not None:
                return 'Material level exist already.', HTTPStatus.CONFLICT

            obj = ResourceMaterialLevel(
                Name=Name,
                LevelNumber=LevelNumber
            )
            obj.save()

            return 'Material level registered successfully.', HTTPStatus.OK

        resourceMaterialLevel = ResourceMaterialLevel.query.filter(ResourceMaterialLevel.Id==Id).first()
        if resourceMaterialLevel:
            resourceMaterialLevel.Name = Name
            resourceMaterialLevel.LevelNumber = LevelNumber
            db.session.commit()

        return 'Material level updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceMaterialLevel')
class deleteResourceMaterialLevel(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource material level
        """
        id = request.args.get('id')
        ResourceMaterialLevel.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource material level deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/uploadTenderResourceManpowerLevel')
class uploadTenderResourceManpowerLevel(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender manpower resource level - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_manpower_resource_level(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        if validation is 200:
            return 'Data Uploaded Successfully!', HTTPStatus.OK

        return 'Internal Server Error!', HTTPStatus.INTERNAL_SERVER_ERROR


@tenders_namespace.route('/uploadTenderResourceManpowerResource')
class uploadTenderResourceManpowerResource(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender manpower resource - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_manpower_resource(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_dict = [v for x, v in validation.items()]
        return {
           'json_dict': json_dict,
           'jsonerror_counter': len(json_dict)
       }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderResourceEquipmentLevel')
class uploadTenderResourceEquipmentLevel(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment resource level - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_equipment_resource_level(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        if validation is 200:
            return 'Data Uploaded Successfully!', HTTPStatus.OK

        return 'Internal Server Error!', HTTPStatus.INTERNAL_SERVER_ERROR


@tenders_namespace.route('/uploadTenderResourceEquipmentResource')
class uploadTenderResourceEquipmentResource(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment resource - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_equipment_resource(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_dict = [v for x,v in validation.items()]
        return {
           'json_dict': json_dict,
           'jsonerror_counter': len(json_dict)
       }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderResourceMaterialLevel')
class uploadTenderResourceMaterialLevel(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender material resource level - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_material_resource_level(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        if validation is 200:
            return 'Data Uploaded Successfully!', HTTPStatus.OK

        return 'Internal Server Error!', HTTPStatus.INTERNAL_SERVER_ERROR


@tenders_namespace.route('/uploadTenderResourceMaterialResource')
class uploadTenderResourceMaterialResource(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender material resource - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'resource_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['resource_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        validation = load_material_resource(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)), get_jwt_identity())

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_dict = [v for x,v in validation.items()]
        return {
           'json_dict': json_dict,
           'jsonerror_counter': len(json_dict)
       }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderManpowerHistogram')
class uploadTenderManpowerHistogram(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender manpower histogram - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = manpower_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            manpower_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 6
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
            'json_dict': json_error,
            'jsonerror_counter': len(json_error)
       }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderManpowerHistogramP6')
class uploadTenderManpowerHistogramP6(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender manpower histogram p6 - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = manpower_p6_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            manpower_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 6
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
            'json_dict': json_error,
            'jsonerror_counter': len(json_error)
       }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderEquipmentHistogram')
class uploadTenderEquipmentHistogram(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment histogram - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = equipment_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            equip_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 7
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
                   'json_dict': json_error,
                   'jsonerror_counter': len(json_error)
        }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderEquipmentHistogramP6')
class uploadTenderEquipmentHistogramP6(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment histogram p6 - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = equipment_p6_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            equip_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 7
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
                   'json_dict': json_error,
                   'jsonerror_counter': len(json_error)
        }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderMaterialHistogram')
class uploadTenderMaterialHistogram(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment histogram - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = material_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            material_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 8
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
                   'json_dict': json_error,
                   'jsonerror_counter': len(json_error)
        }, HTTPStatus.OK


@tenders_namespace.route('/uploadTenderMaterialHistogramP6')
class uploadTenderMaterialHistogramP6(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender equipment histogram p6 - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        json_error, json_dict = material_p6_hist_json_validation(os.path.join(UPLOAD_PATH, 'tenderReourcesExcelFiles/{}'.format(filename)))
        if len(json_error) == 0:
            material_hist_json_to_db(json_dict, tender_id, get_jwt_identity())

            tenderActivityRegister = TenderActivityRegister.query.filter(
                TenderActivityRegister.ten_TenderId == tender_id,
                TenderActivityRegister.ten_ActivityListId == 8
            ).first()

            if tenderActivityRegister:
                tenderActivityRegister.Status = 1
                db.session.commit()

        if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename))):
            os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReourcesExcelFiles', filename)))

        return {
                   'json_dict': json_error,
                   'jsonerror_counter': len(json_error)
        }, HTTPStatus.OK


@tenders_namespace.route('/getManpowerHistogramData')
class getManpowerHistogramData(Resource):

    @jwt_required()
    def get(self):
        """
            get manpower histogram data
        """
        tender_id = request.args.get('tender_id')

        manpowerHistogram = ManpowerHistogram.query.filter(
            ManpowerHistogram.ten_TendersId == tender_id
        ).order_by(ManpowerHistogram.DateStart).all()

        resourceManpower = ResourceManpower.query.all()
        resourceManpower = {x.Id:x.serialize for x in resourceManpower}

        colors = [
            '#FFD100', '#0055B7', '#E2231A', '#000000', '#8b8b8b', '#32b7d2', '#AFC38E', '#e21ac1', '#976865',
            '#1ae2c1', '#e28b1a', '#76c45e', '#a7a7a7'
        ]

        pivot_table = []
        traces = {}
        annotations = {}
        color_i = 0

        resourceManpowerLevel = ResourceManpowerLevel.query.all()
        resourceManpowerLevel = {str(x.Id): x.Name for x in resourceManpowerLevel}

        resourceDepartment = ResourceDepartment.query.all()
        resourceDepartment = {str(x.Id): x.Name for x in resourceDepartment}

        resourceCode = ResourceCode.query.all()
        resourceCode = {str(x.Id): x.Name for x in resourceCode}

        for histogram in manpowerHistogram:
            if resourceManpower[histogram.ten_ResourceManpowerId]['Name'] not in traces:
                traces[resourceManpower[histogram.ten_ResourceManpowerId]['Name']] = {
                    'x': [],
                    'y': [],
                    'name': resourceManpower[histogram.ten_ResourceManpowerId]['Name'],
                    'type': 'bar',
                    'hovertemplate': '<b>{0}</b>: {1}<extra></extra>'.format(resourceManpower[histogram.ten_ResourceManpowerId]['Name'], '%{y}'),
                    'color': colors[color_i]
                }
                color_i += 1
                if len(colors) == color_i:
                    color_i = 0

            traces[resourceManpower[histogram.ten_ResourceManpowerId]['Name']]['x'].append(histogram.DateStart.strftime('%Y-%m'))
            traces[resourceManpower[histogram.ten_ResourceManpowerId]['Name']]['y'].append(round(histogram.Value, 2))

            if histogram.DateStart.strftime('%Y-%m') not in annotations:
                annotations[histogram.DateStart.strftime('%Y-%m')] = 0
            annotations[histogram.DateStart.strftime('%Y-%m')] += histogram.Value

            try:
                level1 = resourceManpowerLevel[resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceManpowerLevelId_1']]
            except:
                level1 = ''

            try:
                level2 = resourceManpowerLevel[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceManpowerLevelId_2']]
            except:
                level2 = ''

            try:
                level3 = resourceManpowerLevel[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceManpowerLevelId_3']]
            except:
                level3 = ''

            try:
                level4 = resourceManpowerLevel[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceManpowerLevelId_4']]
            except:
                level4 = ''

            try:
                level5 = resourceManpowerLevel[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceManpowerLevelId_5']]
            except:
                level5 = ''


            try:
                department = resourceDepartment[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceDepartmentId']]
            except:
                department = ''

            try:
                code = resourceCode[
                    resourceManpower[histogram.ten_ResourceManpowerId]['ten_ResourceCodeId']]
            except:
                code = ''

            pivot_table.append(
                {
                    'Month': histogram.DateStart.strftime('%Y-%m'),
                    'Resource Manpower': resourceManpower[histogram.ten_ResourceManpowerId]['Name'],
                    'Value': round(histogram.Value, 2),
                    'Resource Manpower Level 1': level1,
                    'Resource Manpower Level 2': level2,
                    'Resource Manpower Level 3': level3,
                    'Resource Manpower Level 4': level4,
                    'Resource Manpower Level 5': level5,
                    'Resource Department': department,
                    'Resource Code': code,
                }
            )

        annotations = {k: v for k, v in sorted(list(annotations.items()))}

        line = {
            'x': [],
            'y': [],
            'name': 'Cumulative',
            # 'type': 'line',
            'hovertemplate': [],
            'type': "scatter",
            'mode': "lines",
            'line': {
                'shape': "linear",
                'smoothing': 1,
                'dash': '#0055b7',
                'width': 2,
            },
            'width': 0.5,
            'marker': {
                'color': '#0055b7',
            },
        }

        total = 0
        updated_annotations = []
        for k, v in annotations.items():

            total += v

            line['x'].append(k)
            line['y'].append(total)
            line['hovertemplate'].append('<b>{0}</b>: {1}<extra></extra>'.format(k, '%{y}'))

            updated_annotations.append(
                {
                    'x': k,
                    'y': round(v * 1.05, 2),
                    'text': round(v, 2),
                    'xanchor': 'center',
                    'yanchor': 'auto',
                    'showarrow': False,
                    'font': {
                        'color': "black",
                        'size': 14
                    }
                }
            )

        traces['Cumulative'] = line

        return {
            'traces': traces,
            'annotations': updated_annotations,
            'pivot_table': pivot_table,
        }


@tenders_namespace.route('/getEquipmentHistogramData')
class getEquipmentHistogramData(Resource):

    @jwt_required()
    def get(self):
        """
            get equipment histogram data
        """
        tender_id = request.args.get('tender_id')

        equipmentHistogram = EquipmentHistogram.query.filter(
            EquipmentHistogram.ten_TendersId == tender_id
        ).all()

        resourceEquipment = ResourceEquipment.query.all()
        resourceEquipment = {x.Id:x.serialize for x in resourceEquipment}

        colors = [
            '#FFD100', '#0055B7', '#E2231A', '#000000', '#8b8b8b', '#32b7d2', '#AFC38E', '#e21ac1', '#976865',
            '#1ae2c1', '#e28b1a', '#76c45e', '#a7a7a7'
        ]

        resourceEquipmentLevel = ResourceEquipmentLevel.query.all()
        resourceEquipmentLevel = {str(x.Id): x.Name for x in resourceEquipmentLevel}

        resourceDepartment = ResourceDepartment.query.all()
        resourceDepartment = {str(x.Id): x.Name for x in resourceDepartment}

        resourceCode = ResourceCode.query.all()
        resourceCode = {str(x.Id): x.Name for x in resourceCode}

        pivot_table = []
        traces = {}
        annotations = {}
        color_i = 0

        for histogram in equipmentHistogram:
            if resourceEquipment[histogram.ten_ResourceEquipmentId]['Name'] not in traces:
                traces[resourceEquipment[histogram.ten_ResourceEquipmentId]['Name']] = {
                    'x': [],
                    'y': [],
                    'name': resourceEquipment[histogram.ten_ResourceEquipmentId]['Name'],
                    'type': 'bar',
                    'hovertemplate': '<b>{0}</b>: {1}<extra></extra>'.format(
                        resourceEquipment[histogram.ten_ResourceEquipmentId]['Name'], '%{y}'),
                    'color': colors[color_i]
                }
                color_i += 1
                if len(colors) == color_i:
                    color_i = 0

            traces[resourceEquipment[histogram.ten_ResourceEquipmentId]['Name']]['x'].append(
                histogram.DateStart.strftime('%Y-%m'))
            traces[resourceEquipment[histogram.ten_ResourceEquipmentId]['Name']]['y'].append(round(histogram.Value, 2))

            if histogram.DateStart.strftime('%Y-%m') not in annotations:
                annotations[histogram.DateStart.strftime('%Y-%m')] = 0
            annotations[histogram.DateStart.strftime('%Y-%m')] += histogram.Value

            try:
                level1 = resourceEquipmentLevel[resourceEquipment[histogram.ten_ResourceEquipmentId]['ten_ResourceEquipmentLevelId_1']]
            except:
                level1 = ''

            try:
                level2 = resourceEquipmentLevel[
                    resourceEquipment[histogram.ten_ResourceEquipmentId]['ten_ResourceEquipmentLevelId_2']]
            except:
                level2 = ''

            try:
                level3 = resourceEquipmentLevel[
                    resourceEquipment[histogram.ten_ResourceEquipmentId]['ten_ResourceEquipmentLevelId_3']]
            except:
                level3 = ''

            try:
                department = resourceDepartment[
                    resourceEquipment[histogram.ten_ResourceEquipmentId]['ten_ResourceDepartmentId']]
            except:
                department = ''

            try:
                code = resourceCode[
                    resourceEquipment[histogram.ten_ResourceEquipmentId]['ten_ResourceCodeId']]
            except:
                code = ''

            pivot_table.append(
                {
                    'Month': histogram.DateStart.strftime('%Y-%m'),
                    'Resource Equipment': resourceEquipment[histogram.ten_ResourceEquipmentId]['Name'],
                    'Value': round(histogram.Value, 2),
                    'Resource Equipment Level 1': level1,
                    'Resource Equipment Level 2': level2,
                    'Resource Equipment Level 3': level3,
                    'Resource Department': department,
                    'Resource Code': code,
                }
            )

        annotations = {k: v for k, v in sorted(list(annotations.items()))}

        line = {
            'x': [],
            'y': [],
            'name': 'Cumulative',
            # 'type': 'line',
            'hovertemplate': [],
            'type': "scatter",
            'mode': "lines",
            'line': {
                'shape': "linear",
                'smoothing': 1,
                'dash': '#0055b7',
                'width': 2,
            },
            'width': 0.5,
            'marker': {
                'color': '#0055b7',
            },
        }

        total = 0
        updated_annotations = []
        for k, v in annotations.items():
            total += v

            line['x'].append(k)
            line['y'].append(total)
            line['hovertemplate'].append('<b>{0}</b>: {1}<extra></extra>'.format(k, '%{y}'))

            updated_annotations.append(
                {
                    'x': k,
                    'y': round(v * 1.05, 2),
                    'text': round(v, 2),
                    'xanchor': 'center',
                    'yanchor': 'auto',
                    'showarrow': False,
                    'font': {
                        'color': "black",
                        'size': 14
                    }
                }
            )

        traces['Cumulative'] = line

        return {
            'traces': traces,
            'annotations': updated_annotations,
            'pivot_table': pivot_table,
        }


@tenders_namespace.route('/getMaterialHistogramData')
class getMaterialHistogramData(Resource):

    @jwt_required()
    def get(self):
        """
            get material histogram data
        """
        tender_id = request.args.get('tender_id')

        materialHistogram = MaterialHistogram.query.filter(
            MaterialHistogram.ten_TendersId == tender_id
        ).all()

        resourceMaterial = ResourceMaterial.query.all()
        resourceMaterial = {x.Id:x.serialize for x in resourceMaterial}

        colors = [
            '#FFD100', '#0055B7', '#E2231A', '#000000', '#8b8b8b', '#32b7d2', '#AFC38E', '#e21ac1', '#976865',
            '#1ae2c1', '#e28b1a', '#76c45e', '#a7a7a7'
        ]

        resourceMaterialLevel = ResourceMaterialLevel.query.all()
        resourceMaterialLevel = {str(x.Id): x.Name for x in resourceMaterialLevel}

        resourceDepartment = ResourceDepartment.query.all()
        resourceDepartment = {str(x.Id): x.Name for x in resourceDepartment}

        resourceCode = ResourceCode.query.all()
        resourceCode = {str(x.Id): x.Name for x in resourceCode}

        pivot_table = []
        traces = {}
        annotations = {}
        color_i = 0

        for histogram in materialHistogram:
            if resourceMaterial[histogram.ten_ResourceMaterialId]['Name'] not in traces:
                traces[resourceMaterial[histogram.ten_ResourceMaterialId]['Name']] = {
                    'x': [],
                    'y': [],
                    'name': resourceMaterial[histogram.ten_ResourceMaterialId]['Name'],
                    'type': 'bar',
                    'hovertemplate': '<b>{0}</b>: {1}<extra></extra>'.format(
                        resourceMaterial[histogram.ten_ResourceMaterialId]['Name'], '%{y}'),
                    'color': colors[color_i]
                }
                color_i += 1
                if len(colors) == color_i:
                    color_i = 0

            traces[resourceMaterial[histogram.ten_ResourceMaterialId]['Name']]['x'].append(
                histogram.DateStart.strftime('%Y-%m'))
            traces[resourceMaterial[histogram.ten_ResourceMaterialId]['Name']]['y'].append(round(histogram.Value, 2))

            if histogram.DateStart.strftime('%Y-%m') not in annotations:
                annotations[histogram.DateStart.strftime('%Y-%m')] = 0
            annotations[histogram.DateStart.strftime('%Y-%m')] += histogram.Value

            try:
                level1 = resourceMaterialLevel[resourceMaterial[histogram.ten_ResourceMaterialId]['ten_ResourceMaterialLevelId_1']]
            except:
                level1 = ''

            try:
                department = resourceDepartment[
                    resourceMaterial[histogram.ten_ResourceMaterialId]['ten_ResourceDepartmentId']]
            except:
                department = ''

            try:
                code = resourceCode[
                    resourceMaterial[histogram.ten_ResourceMaterialId]['ten_ResourceCodeId']]
            except:
                code = ''

            pivot_table.append(
                {
                    'Month': histogram.DateStart.strftime('%Y-%m'),
                    'Resource Material': resourceMaterial[histogram.ten_ResourceMaterialId]['Name'],
                    'Value': round(histogram.Value, 2),
                    'Resource Material Level 1': level1,
                    'Resource Department': department,
                    'Resource Code': code,
                }
            )

        annotations = {k: v for k, v in sorted(list(annotations.items()))}

        line = {
            'x': [],
            'y': [],
            'name': 'Cumulative',
            # 'type': 'line',
            'hovertemplate': [],
            'type': "scatter",
            'mode': "lines",
            'line': {
                'shape': "linear",
                'smoothing': 1,
                'dash': '#0055b7',
                'width': 2,
            },
            'width': 0.5,
            'marker': {
                'color': '#0055b7',
            },
        }

        total = 0
        updated_annotations = []
        for k, v in annotations.items():
            total += v

            line['x'].append(k)
            line['y'].append(round(total, 2))
            line['hovertemplate'].append('<b>{0}</b>: {1}<extra></extra>'.format(k, '%{y}'))

            updated_annotations.append(
                {
                    'x': k,
                    'y': round(v * 1.05, 2),
                    'text': round(v, 2),
                    'xanchor': 'center',
                    'yanchor': 'auto',
                    'showarrow': False,
                    'font': {
                        'color': "black",
                        'size': 14
                    }
                }
            )

        traces['Cumulative'] = line

        return {
            'traces': traces,
            'annotations': updated_annotations,
            'pivot_table': pivot_table,
        }


@tenders_namespace.route('/updateResourceCode')
class updateResourceCode(Resource):

    @jwt_required()
    def post(self):
        """
            update resource code
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']

        if Id == '0':
            resourceCode = ResourceCode.query.filter(ResourceCode.Name == Name).first()
            if resourceCode is not None:
                return 'Resource code exist already.', HTTPStatus.CONFLICT

            obj = ResourceCode(
                Name=Name,
            )
            obj.save()

            return 'Resource code registered successfully.', HTTPStatus.OK

        resourceCode = ResourceCode.query.filter(ResourceCode.Id==Id).first()
        if resourceCode:
            resourceCode.Name = Name
            db.session.commit()

        return 'Resource code updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceCode')
class deleteResourceCode(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource code
        """
        id = request.args.get('id')
        ResourceCode.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource code deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/updateResourceDepartment')
class updateResourceDepartment(Resource):

    @jwt_required()
    def post(self):
        """
            update resource department
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']

        if Id == '0':
            resourceDepartment = ResourceDepartment.query.filter(ResourceDepartment.Name == Name).first()
            if resourceDepartment is not None:
                return 'Resource department exist already.', HTTPStatus.CONFLICT

            obj = ResourceDepartment(
                Name=Name,
            )
            obj.save()

            return 'Resource department registered successfully.', HTTPStatus.OK

        resourceDepartment = ResourceDepartment.query.filter(ResourceDepartment.Id==Id).first()
        if resourceDepartment:
            resourceDepartment.Name = Name
            db.session.commit()

        return 'Resource department updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteResourceDepartment')
class deleteResourceDepartment(Resource):

    @jwt_required()
    def get(self):
        """
            delete resource department
        """
        id = request.args.get('id')
        ResourceDepartment.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Resource department deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/updateActivity')
class updateActivity(Resource):

    @jwt_required()
    def post(self):
        """
            add/update activity
        """
        data = request.get_json()

        Id = data['Id']
        Name = data['Name']
        OrderNumber = data['OrderNumber']
        PositionId = data['PositionId']

        if Id == '0':
            activityList = ActivityList.query.filter(
                ActivityList.Name == Name,
            ).first()
            if activityList is not None:
                return 'Activity exist already.', HTTPStatus.CONFLICT

            obj = ActivityList(
                Name=Name,
                OrderNumber=OrderNumber,
                PositionId=PositionId,
                CreatedOn=datetime.now(),
                CreatedBy=get_jwt_identity(),
            )
            obj.save()

            return 'Activity registered successfully.', HTTPStatus.OK

        activityList = ActivityList.query.filter(ActivityList.Id==Id).first()
        if activityList:
            activityList.Name = Name
            activityList.OrderNumber = OrderNumber
            activityList.PositionId = PositionId
            db.session.commit()

        return 'Activity updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteActivity')
class deleteActivity(Resource):

    @jwt_required()
    def get(self):
        """
            delete Activity
        """
        activity_id = request.args.get('id')
        ActivityList.query.filter_by(Id=activity_id).delete()
        db.session.commit()
        return 'Activity deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/triggerDeadlineEmail')
class triggerDeadlineEmail(Resource):

    @jwt_required()
    def get(self):

        tenderActivityRegister = TenderActivityRegister.query.filter(
            TenderActivityRegister.emailTriggered == False,
            TenderActivityRegister.Deadline != None,
            or_(
                TenderActivityRegister.Status == 0,
                TenderActivityRegister.Status == None,
            ),
            TenderActivityRegister.ten_UserListId != None,
        ).all()

        for tenderActivity in tenderActivityRegister:
            deadline = tenderActivity.Deadline.date()
            today = datetime.now().date()

            days = (deadline - today).days
            # days = 2
            if days > 0 and days < 3:

                user = UserList.query.filter(
                    UserList.Id == tenderActivity.ten_UserListId
                ).first()

                if user:
                    # user.dbo_AspNetUsersId = 'aa1493db-36c6-48a8-ba90-51711bcef2e4'
                    User_Obj = User.query.filter(User.id == user.dbo_AspNetUsersId).first()

                    if User_Obj and User_Obj.email:
                        mail_body = r'Hi, <br><br>' \
                                    r'The deadline is near, please go to tendering app for further actions.<br>' \
                                    r'<br>NMDC Webapps,'

                        port = 587  # For starttls
                        smtp_server = "mail.nmdc.ae"
                        sender_email = "webapps@nmdc.ae"
                        receiver_email = [User_Obj.email]

                        message = MIMEMultipart("alternative")
                        message["Subject"] = 'Tenders - Deadline'
                        message["From"] = sender_email
                        message["To"] = ", ".join(receiver_email)

                        body = MIMEText(mail_body, 'html')
                        message.attach(body)

                        with smtplib.SMTP(smtp_server, port) as server:
                            server.sendmail(sender_email, receiver_email, message.as_string())

                        tenderActivity.emailTriggered = 1
                        db.session.commit()

        return 'Email Sent Successfully.', HTTPStatus.OK


@tenders_namespace.route('/getAnalysisExcel')
class getAnalysisExcel(Resource):

    def get(self):
        """
            get Analysis Excel
        """
        # start_date = request.args.get('start_date')
        # end_date = request.args.get('end_date')


        foldername = 'tenderReports'
        filename = 'excel_file.xlsx'

        # code to create excel file and save in 'tenderReports' folder

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        uploads = os.path.join(BASE_DIR, 'static\\' + foldername)

        if os.path.exists(os.path.join(uploads, filename)):
            response = send_from_directory(directory=uploads, path=filename, as_attachment=True)
            response.headers['Content-Type'] = 'application/vnd'
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            return response
        else:
            return 'File not found.', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/getWorkloadExcel')
class getWorkloadExcel(Resource):

    def get(self):
        """
            get workload loadExcel
        """
        # tender_id = request.args.get('tender_id').split(',')
        start_date = datetime.strptime(request.args.get('start_date'), '%m/%d/%Y').date()
        end_date = datetime.strptime(request.args.get('end_date'), '%m/%d/%Y').date()
        # tender_obj = Tenders.query.filter(Tenders.Id.in_(tender_id)).all()
        tender_obj = Tenders.query.filter(and_(Tenders.SubmissionDate>=start_date,\
                                                Tenders.SubmissionDate<=end_date,\
                                                Tenders.Submission=='Main')).all()

        status_obj = TenderStatus.query.all()
        status_dict = {}
        for st_ins in status_obj:
            status_dict[st_ins.Id] = st_ins.TenderStatusName

        wb = openpyxl.Workbook()
        work_sheet = wb.create_sheet('Workload')

        c1 = work_sheet.cell(row=1, column=1, value='Tender Code')
        c1 = work_sheet.cell(row=1, column=2, value='Project')
        c1 = work_sheet.cell(row=1, column=3, value='Location')
        c1 = work_sheet.cell(row=1, column=4, value='Client')
        c1 = work_sheet.cell(row=1, column=5, value='Submission Date')
        c1 = work_sheet.cell(row=1, column=6, value='Scope of Works')
        c1 = work_sheet.cell(row=1, column=7, value='Civil Marine / EPC')
        c1 = work_sheet.cell(row=1, column=8, value='GI & Piling Works')
        c1 = work_sheet.cell(row=1, column=9, value='Dredging')
        c1 = work_sheet.cell(row=1, column=10, value='Total')
        c1 = work_sheet.cell(row=1, column=11, value='Probability')
        c1 = work_sheet.cell(row=1, column=12, value='remarks')
        c1 = work_sheet.cell(row=1, column=13, value='TenderStatus')

        prev_row = 2
        for tender_ins in tender_obj:
            ten_submission_list = tender_ins.tenderSubmissionDates
            submission = None
            for sub_date in ten_submission_list:
                if submission is None:
                    submission = sub_date.SubmissionDate
                else:
                    if sub_date.SubmissionDate > submission:
                        submission = sub_date.SubmissionDate

            # if submission >= start_date and submission <= end_date:

            c1 = work_sheet.cell(row=prev_row, column=1, value=tender_ins.TenderNumber)
            c1 = work_sheet.cell(row=prev_row, column=2, value=tender_ins.ProjectName)
            c1 = work_sheet.cell(row=prev_row, column=3, value=tender_ins.Place)
            c1 = work_sheet.cell(row=prev_row, column=4, value=tender_ins.Employer_Id)
            c1 = work_sheet.cell(row=prev_row, column=5, value=submission)
            c1 = work_sheet.cell(row=prev_row, column=6, value=tender_ins.ScopeNarrative)
            c1 = work_sheet.cell(row=prev_row, column=7, value=tender_ins.CivilMarineEPC)
            c1 = work_sheet.cell(row=prev_row, column=8, value=tender_ins.GIPillingWorks)
            c1 = work_sheet.cell(row=prev_row, column=9, value=tender_ins.Dredging)
            total = 0
            if tender_ins.CivilMarineEPC is not None and tender_ins.CivilMarineEPC > 0:
                total += tender_ins.CivilMarineEPC
            if tender_ins.GIPillingWorks is not None and tender_ins.GIPillingWorks > 0:
                total += tender_ins.GIPillingWorks
            if tender_ins.Dredging is not None and tender_ins.Dredging > 0:
                total += tender_ins.Dredging
            c1 = work_sheet.cell(row=prev_row, column=10, value=total)
            c1 = work_sheet.cell(row=prev_row, column=11, value=tender_ins.Probability)
            c1 = work_sheet.cell(row=prev_row, column=12, value=tender_ins.Remarks)
            tender_status = None
            if tender_ins.TenderStatus_Id in status_dict:
                tender_status = status_dict[tender_ins.TenderStatus_Id]
            c1 = work_sheet.cell(row=prev_row, column=13, value=tender_status)
            print(tender_ins.TenderNumber, tender_ins.ProjectName, tender_ins.Place, submission, tender_ins.Employer_Id,
                    tender_ins.ScopeNarrative, tender_ins.CivilMarineEPC, tender_ins.GIPillingWorks, tender_ins.Dredging,
                    tender_ins.Remarks)
            prev_row += 1

        UPLOAD_PATH = config('UPLOAD_PATH')
        foldername = 'tenderReports'
        filename = 'workload.xlsx'

        wb.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderReports', filename)))

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        uploads = os.path.join(BASE_DIR, 'static\\' + foldername)

        if os.path.exists(os.path.join(uploads, filename)):
            response = send_from_directory(directory=uploads, path=filename, as_attachment=True)
            response.headers['Content-Type'] = 'application/vnd'
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            return response
        else:
            return 'File not found.', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/uploadTenderFinancialIneightData')
class uploadTenderFinancialIneightData(Resource):

    @jwt_required()
    def post(self):
        """
            upload tender financial ineight data - Excel file
        """
        UPLOAD_PATH = config('UPLOAD_PATH')

        if 'excel_data' not in request.files:
            return 'No File in Content', HTTPStatus.NO_CONTENT

        tender_id = request.form['tender_id']
        file = request.files['excel_data']

        if file.filename == '':
            return 'File Not Found', HTTPStatus.NOT_FOUND

        filename = secure_filename(file.filename)
        filename, file_extension = os.path.splitext(filename)

        filename = ''.join(e for e in filename if e.isalnum())
        filename = str(tender_id) + '_' + datetime.now().strftime("%m%d%Y_%H%M%S") + filename + file_extension
        file.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderFinancialIneightFiles', filename)))

        tender = Tenders.query.filter(
            Tenders.Id == tender_id
        ).first()

        if tender:
            if os.path.exists(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderFinancialIneightFiles', tender.In8FileName))):
                os.remove(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderFinancialIneightFiles', tender.In8FileName)))

            tender.In8FileName = filename
            db.session.commit()

        in8_data_path = os.path.join(UPLOAD_PATH, 'tenderFinancialIneightFiles/{}'.format(filename))
        new_dataform_path = os.path.join(UPLOAD_PATH, 'tenderFinancialIneightFiles/{}'.format('Tender_Data_Form_New_sample.xlsx'))

        json_error, dataf_load_json = in8_to_dataform_val(in8_data_path, new_dataform_path)
        jsonerror_counter = len(json_error.keys())

        filename = ''
        if jsonerror_counter == 0:
            filename = in8_to_dataform_gen(dataf_load_json, new_dataform_path)
            tenders = Tenders.query.filter(Tenders.Id==tender_id).first()
            tenders.FinancialInEightFileLink = filename
            db.session.commit()

        return {
            'json_dict': json_error,
            'jsonerror_counter': jsonerror_counter,
            'fileName': filename
        }, HTTPStatus.OK


def in8_to_dataform_gen(dataf_load_json, sample_path):
    gen_df_obj = openpyxl.load_workbook(filename=sample_path, data_only=False)
    gen_df_sheet = gen_df_obj['DataForm']

    for r in dataf_load_json:
        for c in dataf_load_json[r]:
            c1 = gen_df_sheet.cell(row = r, column=c,value =dataf_load_json[r][c] )

    UPLOAD_PATH = config('UPLOAD_PATH')
    filename = datetime.now().strftime("%m%d%Y_%H%M%S") + 'Tender_Data_Form_New_Generated.xlsx'
    gen_df_obj.save(os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderFinancialIneightFiles', filename)))

    return filename


@tenders_namespace.route('/getOrganizationalCategory')
class getOrganizationalCategory(Resource):

    @jwt_required()
    def get(self):
        """
            get Organizational Category
        """

        to_remove = ['Id', 'ten_TenderId', 'ten_EquipmentCostItemId', 'IsArchived', 'CreatedBy', 'CreatedOn']
        columns = TenderFinancialData.__table__.columns.keys()
        columns = list(filter(lambda item: item not in to_remove, columns))

        organizationalCategory = OrganizationalCategory.query.all()

        return {
            'columns': columns,
            'organizationalCategory': [x.serialize for x in organizationalCategory]
        }, HTTPStatus.OK


@tenders_namespace.route('/updateOrganizationalCategory')
class updateOrganizationalCategory(Resource):

    @jwt_required()
    def post(self):
        """
            add/update Organizational Category
        """
        data = request.get_json()

        Id = data['Id']
        OrganizationalCategoryItem = data['OrganizationalCategoryItem']
        ExcelColumnName = data['ExcelColumnName']
        TopSheetDescRef = data['TopSheetRefName']
        TopSheetItemCode = data['TopSheetItemCode']

        if Id == '0':
            organizationalCategory = OrganizationalCategory.query.filter(
                OrganizationalCategory.OrganizationalCategoryItem == OrganizationalCategoryItem,
            ).first()
            if organizationalCategory is not None:
                return 'Organizational Category exist already.', HTTPStatus.CONFLICT

            obj = OrganizationalCategory(
                OrganizationalCategoryItem=OrganizationalCategoryItem,
                ExcelColumnName=ExcelColumnName,
                TopSheetDescRef=TopSheetDescRef,
                TopSheetItemCode=TopSheetItemCode,
            )
            obj.save()

            return 'Organizational Category registered successfully.', HTTPStatus.OK

        organizationalCategory = OrganizationalCategory.query.filter(OrganizationalCategory.Id==Id).first()
        if organizationalCategory:
            organizationalCategory.OrganizationalCategoryItem = OrganizationalCategoryItem
            organizationalCategory.ExcelColumnName = ExcelColumnName
            organizationalCategory.TopSheetDescRef = TopSheetDescRef
            organizationalCategory.TopSheetItemCode = TopSheetItemCode
            db.session.commit()

        return 'Organizational Category updated successfully.', HTTPStatus.OK


@tenders_namespace.route('/deleteOrganizationalCategory')
class deleteOrganizationalCategory(Resource):

    @jwt_required()
    def get(self):
        """
            delete Organizational Category
        """
        id = request.args.get('id')
        OrganizationalCategory.query.filter_by(Id=id).delete()
        db.session.commit()
        return 'Organizational Category deleted successfully.', HTTPStatus.OK


@tenders_namespace.route('/getTopSheet')
class getTopSheet(Resource):

    def get(self):
        """
            get tender topsheet excel
        """
        tender_id = request.args.get('tender_id')

        UPLOAD_PATH = config('UPLOAD_PATH')
        delete_files_from_folder(os.path.join(UPLOAD_PATH, 'tenderTopsheet'))

        filename = 'TopsheetSample.xlsx'

        topsheet_path = os.path.join(UPLOAD_PATH, '{}/{}'.format('tenderFinancialIneightFiles', filename))
        org_obj = OrganizationalCategory.query.filter(OrganizationalCategory.TopSheetItemCode != None).all()
        org_dict = {}
        for org_ins in org_obj:
            org_dict[org_ins.ExcelColumnName] = {'Id': org_ins.Id, 'Name': org_ins.TopSheetDescRef,
                                                 'Code': org_ins.TopSheetItemCode}

        eq_obj = EquipmentCostItem.query.filter(EquipmentCostItem.ParentId == None).all()
        eq_dict = []
        for eq_ins in eq_obj:
            eq_dict.append(eq_ins.Id)

        fin_obj = TenderFinancialData.query.filter(and_(TenderFinancialData.ten_TenderId == tender_id,
                                                        TenderFinancialData.ten_EquipmentCostItemId.in_(eq_dict))).all()
        top_code_dict = {}
        if len(fin_obj) > 0:
            for fin_ins in fin_obj:
                for org_i in org_dict:
                    if org_dict[org_i]['Code'] not in top_code_dict:
                        top_code_dict[org_dict[org_i]['Code']] = getattr(fin_ins, org_i)
                    else:
                        top_code_dict[org_dict[org_i]['Code']] += getattr(fin_ins, org_i)

        gen_df_obj = openpyxl.load_workbook(filename=topsheet_path, data_only=False)
        gen_df_sheet = gen_df_obj['T_sheet_gen']

        top_sheet_row_d = {}
        for r_i in gen_df_sheet['A13:A121']:
            if r_i[0].value is not None:
                if r_i[0].value not in top_sheet_row_d:
                    top_sheet_row_d[r_i[0].value] = r_i[0].row

        for t_ins in top_code_dict:
            gen_df_sheet.cell(row=top_sheet_row_d[t_ins], column=5, value=top_code_dict[t_ins])

        foldername = 'tenderTopsheet'

        gen_df_obj.save(os.path.join(UPLOAD_PATH, '{}/{}'.format(foldername, filename)))

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
        uploads = os.path.join(BASE_DIR, 'static\\' + foldername)

        if os.path.exists(os.path.join(uploads, filename)):
            response = send_from_directory(directory=uploads, path=filename, as_attachment=True)
            response.headers['Content-Type'] = 'application/vnd'
            response.headers['Content-Disposition'] = 'attachment; filename={0}'.format(filename)
            return response
        else:
            return 'File not found.', HTTPStatus.NOT_FOUND


@tenders_namespace.route('/getTenderAnalysisTable')
class getTenderAnalysisTable(Resource):

    @jwt_required()
    def get(self):

        tenders = Tenders.query.filter(
            or_(
                Tenders.IsArchive == None,
                Tenders.IsArchive == False,
            )
        ).all()

        response_data = []
        for tender in tenders:

            # if len(tender.tenderActivityRegister) == 0:
            #
            #     activityList = ActivityList.query.order_by(ActivityList.OrderNumber).all()
            #     for activity in activityList:
            #
            #         Status = False
            #         ten_UserListId = None
            #         if activity.OrderNumber == 1:
            #             Status = True
            #
            #         if activity.OrderNumber == 3 and len(tender.tenderFinancialData) > 0:
            #             Status = True
            #
            #         if activity.OrderNumber == 4 and len(tender.tenderEquipmentCost) > 0:
            #             Status = True
            #
            #         if activity.OrderNumber == 5 and len(tender.planning) > 0:
            #             Status = True
            #
            #         if activity.OrderNumber == 6:
            #             manpowerHistogram = ManpowerHistogram.query.filter(
            #                 ManpowerHistogram.ten_TendersId == tender.Id
            #             ).all()
            #
            #             if len(manpowerHistogram) > 0:
            #                 Status = True
            #
            #         if activity.OrderNumber == 7:
            #             equipmentHistogram = EquipmentHistogram.query.filter(
            #                 EquipmentHistogram.ten_TendersId == tender.Id
            #             ).all()
            #
            #             if len(equipmentHistogram) > 0:
            #                 Status = True
            #
            #         if activity.OrderNumber == 8:
            #             materialHistogram = MaterialHistogram.query.filter(
            #                 MaterialHistogram.ten_TendersId == tender.Id
            #             ).all()
            #
            #             if len(materialHistogram) > 0:
            #                 Status = True
            #
            #         if activity.OrderNumber == 9 and len(tender.riskRegister) > 0:
            #             Status = True
            #
            #         tenderActivityRegister = TenderActivityRegister(
            #             ten_TenderId=tender.Id,
            #             ten_ActivityListId=activity.OrderNumber,
            #             ten_UserListId=ten_UserListId,
            #             Deadline=None,
            #             Status=Status,
            #             CreatedBy=get_jwt_identity(),
            #             CreatedOn=datetime.now(),
            #             UpdatedBy=get_jwt_identity(),
            #             UpdatedOn=datetime.now(),
            #         )
            #         tenderActivityRegister.save()
            #
            #         if activity.OrderNumber == 2:
            #             updated_tick_status(tender, tender.Id)

            try:
                Info_Contracts = next(filter(lambda x: x.ten_ActivityListId == 2, tender.tenderActivityRegister)).Status
            except:
                Info_Contracts = False

                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=2,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=False,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()
                # updated_tick_status(tender, tender.Id)

            try:
                Financial_Cost = next(filter(lambda x: x.ten_ActivityListId == 3, tender.tenderActivityRegister)).Status
            except:
                Financial_Cost = False

                # Status = False
                # if len(tender.tenderFinancialData) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=3,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Equipment_Cost = next(filter(lambda x: x.ten_ActivityListId == 4, tender.tenderActivityRegister)).Status
            except:
                Equipment_Cost = False

                # Status = False
                # if len(tender.tenderEquipmentCost) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=4,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Planning = next(filter(lambda x: x.ten_ActivityListId == 5, tender.tenderActivityRegister)).Status
            except:
                Planning = False

                # Status = False
                # if len(tender.planning) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=5,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Manpower_Histogram = next(filter(lambda x: x.ten_ActivityListId == 6, tender.tenderActivityRegister)).Status
            except:
                Manpower_Histogram = False

                # Status = False
                # manpowerHistogram = ManpowerHistogram.query.filter(
                #     ManpowerHistogram.ten_TendersId == tender.Id
                # ).all()
                #
                # if len(manpowerHistogram) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=6,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Equipment_Histogram = next(filter(lambda x: x.ten_ActivityListId == 7, tender.tenderActivityRegister)).Status
            except:
                Equipment_Histogram = False

                # Status = False
                # equipmentHistogram = EquipmentHistogram.query.filter(
                #     EquipmentHistogram.ten_TendersId == tender.Id
                # ).all()
                #
                # if len(equipmentHistogram) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=7,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Material_Histogram = next(filter(lambda x: x.ten_ActivityListId == 8, tender.tenderActivityRegister)).Status
            except:
                Material_Histogram = False

                # Status = False
                # materialHistogram = MaterialHistogram.query.filter(
                #     MaterialHistogram.ten_TendersId == tender.Id
                # ).all()
                #
                # if len(materialHistogram) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=8,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            try:
                Risk_Register = next(filter(lambda x: x.ten_ActivityListId == 9, tender.tenderActivityRegister)).Status
            except:
                Risk_Register = False

                # Status = False
                # if len(tender.riskRegister) > 0:
                #     Status = True
                #
                # tenderActivityRegister = TenderActivityRegister(
                #     ten_TenderId=tender.Id,
                #     ten_ActivityListId=9,
                #     ten_UserListId=None,
                #     Deadline=None,
                #     Status=Status,
                #     CreatedBy=get_jwt_identity(),
                #     CreatedOn=datetime.now(),
                #     UpdatedBy=get_jwt_identity(),
                #     UpdatedOn=datetime.now(),
                # )
                # tenderActivityRegister.save()

            submission_dates = []
            for date in tender.tenderSubmissionDates:
                submission_dates.append(date.SubmissionDate)

            try:
                date = max(submission_dates).strftime('%Y')
            except:
                date = ''

            if date == '1900':
                date = ''

            response_data.append(
                {
                    'Tender_Number': tender.TenderNumber,
                    'Project_Name': tender.ProjectName,
                    'Submission_Date': date,
                    'TenderSplitType': tender.TenderSplitType,
                    'Info_Contracts': Info_Contracts,
                    'Financial_Cost': Financial_Cost,
                    'Equipment_Cost': Equipment_Cost,
                    'Planning': Planning,
                    'Manpower_Histogram': Manpower_Histogram,
                    'Equipment_Histogram': Equipment_Histogram,
                    'Material_Histogram': Material_Histogram,
                    'Risk_Register': Risk_Register,
                }
            )



        return response_data, HTTPStatus.OK