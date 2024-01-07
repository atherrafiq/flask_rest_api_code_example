import openpyxl
from datetime import datetime

from sqlalchemy import or_

from ..models.tenders_models import *


class LoadTenderDataForm():
    def __init__(self, path):
        ### Set max font family value to 100 ###
        self.wb_obj = openpyxl.load_workbook(filename=path, data_only=True)
        self.dataform_sheet = self.wb_obj['DataForm']

        self.createdby = 'excel_import_pkr'
        self.createdon = datetime.now()
        self.json_dict = {'item_name': {}}
        self.json_error = {}
        self.jsonerror_counter = 0

        self.fin_data_ref_dict = {'EquipmentEME': 3,
                                  'EquipmentInternal': 4,
                                  'EquipmentExternal': 5,
                                  'Fuel': 6,
                                  'MaterialsPipeline': 7,
                                  'MaterialsTeeth': 8,
                                  'MaterialsSteel': 9,
                                  'MaterialsRock': 10,
                                  'MaterialsImportedFill': 11,
                                  'MaterialsGeotextile': 12,
                                  'MaterialsConcrete': 13,
                                  'MaterialsNavAid': 14,
                                  'MaterialsMarineFurniture': 15,
                                  'MaterialsOther': 16,
                                  'PersonalLabour': 18,
                                  'PersonalStaff': 19,
                                  'Misc': 20,
                                  'Subcontractor': 21,
                                  'Contingencies': 23,
                                  'MarkupsRisk': 25,
                                  'MarkupsOverhead': 26,
                                  'MarkupsProfit': 27,
                                  'TaxDirect': 29,
                                  'TaxIndirect': 30,
                                  'Quantity': 34,
                                  'Unit': 33,
                                  'Remarks': 37}

        equipment_cost_obj = EquipmentCostItem.query.all()
        self.equipment_cost_dict = {}
        for eq in equipment_cost_obj:
            self.equipment_cost_dict[(eq.Item).lower().replace(' ', '')] = eq.Id

    def financial_data(self):
        for dat in self.dataform_sheet['B10:b77']:
            if dat[0].value is not None:
                item_name_lw = (dat[0].value).lower().replace(' ', '')
                cur_row_idx = dat[0].row

                # Check the item name in db
                if item_name_lw in self.equipment_cost_dict:
                    self.json_dict['item_name'][item_name_lw] = {'id': self.equipment_cost_dict[item_name_lw],
                                                                 'status': 1}
                    self.json_dict[dat[0].value] = {}

                    ## Checking column data for each row
                    for col_name in self.fin_data_ref_dict:
                        cur_col_idx = self.fin_data_ref_dict[col_name]
                        cell_obj = self.dataform_sheet.cell(row=cur_row_idx, column=cur_col_idx)

                        ## first cond checks for column with numeric data
                        if col_name not in ['Unit', 'Remarks']:
                            if cell_obj.value is not None:
                                if isinstance(cell_obj.value, float) or isinstance(cell_obj.value, int):
                                    self.json_dict[dat[0].value][col_name] = {'value': cell_obj.value, 'status': 1}
                                else:
                                    self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 0,
                                                                              'remarks': f'Data for {dat[0].value} under {col_name} is not numeric'}
                                    self.jsonerror_counter += 1
                                    self.json_error[
                                        self.jsonerror_counter] = f'Data for {dat[0].value} under {col_name} is not numeric'
                            else:
                                self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 1}

                        ## Condition to check string data type for the unit and remarks column
                        elif col_name in ['Unit', 'Remarks']:
                            if cell_obj.value is not None:
                                if isinstance(cell_obj.value, str):
                                    self.json_dict[dat[0].value][col_name] = {'value': cell_obj.value, 'status': 1}
                                else:
                                    self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 0,
                                                                              'remarks': f'Data for {dat[0].value} under {col_name} is not string'}
                                    self.jsonerror_counter += 1
                                    self.json_error[
                                        self.jsonerror_counter] = f'Data for {dat[0].value} under {col_name} is not string'
                            else:
                                self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 1}

                else:
                    self.json_dict['item_name'] = {'id': self.equipment_cost_dict[item_name_lw], 'status': 0,
                                                   'remarks': f'Item name not found in database. Check cell {dat[0].coordinates}'}
                    self.jsonerror_counter += 1
                    self.json_error[
                        self.jsonerror_counter] = f'Item name not found in database. Check cell {dat[0].coordinates}'


class new_LoadTenderDataForm():
    def __init__(self, path):
        ### Set max font family value to 100 ###
        self.wb_obj = openpyxl.load_workbook(filename=path, data_only=True)
        self.dataform_sheet = self.wb_obj['DataForm']

        self.createdby = 'excel_import_pkr'
        self.createdon = datetime.now()
        self.json_dict = {'item_name': {}}
        self.json_error = {}
        self.jsonerror_counter = 0

        self.fin_data_ref_dict = {'EquipmentEME': 3,
                                  'EquipmentInternal': 4,
                                  'EquipmentExternal': 5,
                                  'Fuel': 6,
                                  'MaterialsPipeline': 7,
                                  'MaterialsTeeth': 8,
                                  'MaterialsSteel': 9,
                                  'MaterialsRock': 10,
                                  'MaterialsImportedFill': 11,
                                  'MaterialsGeotextile': 12,
                                  'MaterialsConcrete': 13,
                                  'MaterialsNavAid': 14,
                                  'MaterialsMarineFurniture': 15,
                                  'MaterialsOther': 16,
                                  'PersonalLabour': 18,
                                  'PersonalStaff': 19,
                                  'Misc': 20,
                                  'Subcontractor': 21,
                                  'Contingencies': 23,
                                  'MarkupsRiskPercent': 25,
                                  'MarkupsOverheadPercent': 26,
                                  'MarkupsProfitPercent': 27,
                                  'TaxDirect': 29,
                                  'TaxIndirect': 30,
                                  'Quantity': 34,
                                  'Unit': 33,
                                  'Remarks': 37}

        equipment_cost_obj = EquipmentCostItem.query.all()
        self.equipment_cost_dict = {}
        equip_dict = {}
        sub_child = []
        child = []
        for equip_ins in equipment_cost_obj:
            self.equipment_cost_dict[(equip_ins.Item).lower().replace(' ', '')] = equip_ins.Id
            equip_dict[equip_ins.Id] = {'name': equip_ins.Item, 'parentid': equip_ins.ParentId}
            if equip_ins.Child is not None:
                child.append(equip_ins.Id)
            if equip_ins.SubChild is not None:
                sub_child.append(equip_ins.Id)

        calc_check_dict = {}
        for s in sub_child:
            if equip_dict[s]['parentid'] not in calc_check_dict:
                calc_check_dict[equip_dict[s]['parentid']] = []
            calc_check_dict[equip_dict[s]['parentid']].append(s)

        for c in child:
            if equip_dict[c]['parentid'] not in calc_check_dict:
                calc_check_dict[equip_dict[c]['parentid']] = []
            calc_check_dict[equip_dict[c]['parentid']].append(c)


        self.calc_check_item_dict = {}
        for cal in calc_check_dict:
            self.calc_check_item_dict[equip_dict[cal]['name']] = []
            for cal_int in calc_check_dict[cal]:
                self.calc_check_item_dict[equip_dict[cal]['name']].append(equip_dict[cal_int]['name'])

    def financial_data(self):
        for dat in self.dataform_sheet['B10:B80']:
            if dat[0].value is not None:
                item_name_lw = (dat[0].value).lower().replace(' ', '')
                cur_row_idx = dat[0].row

                # Check the item name in db
                if item_name_lw in self.equipment_cost_dict:
                    self.json_dict['item_name'][item_name_lw] = {'id': self.equipment_cost_dict[item_name_lw],
                                                                 'status': 1}
                    self.json_dict[dat[0].value] = {}

                    ## Checking column data for each row
                    for col_name in self.fin_data_ref_dict:
                        cur_col_idx = self.fin_data_ref_dict[col_name]
                        cell_obj = self.dataform_sheet.cell(row=cur_row_idx, column=cur_col_idx)

                        ## first cond checks for column with numeric data
                        if col_name not in ['Unit', 'Remarks']:
                            if cell_obj.value is not None:
                                if isinstance(cell_obj.value, float) or isinstance(cell_obj.value, int):
                                    self.json_dict[dat[0].value][col_name] = {'value': cell_obj.value, 'status': 1}
                                else:
                                    self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 0,
                                                                              'remarks': f'Data for {dat[0].value} under {col_name} is not numeric'}
                                    self.jsonerror_counter += 1
                                    self.json_error[
                                        self.jsonerror_counter] = f'Data for {dat[0].value} under {col_name} is not numeric'
                            else:
                                # print("old",dat[0].value,col_name,cell_obj.coordinate)
                                self.json_dict[dat[0].value][col_name] = {'value': 0, 'status': 1}

                        ## Condition to check string data type for the unit and remarks column
                        elif col_name in ['Unit', 'Remarks']:
                            if cell_obj.value is not None:
                                if isinstance(cell_obj.value, str):
                                    self.json_dict[dat[0].value][col_name] = {'value': cell_obj.value, 'status': 1}
                                else:
                                    self.json_dict[dat[0].value][col_name] = {'value': '', 'status': 0,
                                                                              'remarks': f'Data for {dat[0].value} under {col_name} is not string'}
                                    self.jsonerror_counter += 1
                                    self.json_error[
                                        self.jsonerror_counter] = f'Data for {dat[0].value} under {col_name} is not string'
                            else:
                                self.json_dict[dat[0].value][col_name] = {'value': '', 'status': 1}

                else:
                    self.json_dict['item_name'] = {'id': self.equipment_cost_dict[item_name_lw], 'status': 0,
                                                   'remarks': f'Item name not found in database. Check cell {dat[0].coordinates}'}
                    self.jsonerror_counter += 1
                    self.json_error[
                        self.jsonerror_counter] = f'Item name not found in database. Check cell {dat[0].coordinates}'
 
        if self.jsonerror_counter == 0:
            ## check the calculation in the financial cost
            for cal in self.calc_check_item_dict:
                for col in self.fin_data_ref_dict:
                    if col not in ['Unit', 'Remarks', 'Quantity', 'MarkupsRiskPercent', 'MarkupsOverheadPercent',
                                   'MarkupsProfitPercent']:
                        parent_val = self.json_dict[cal][col]['value']
                        child_val = 0
                        for child_cal in self.calc_check_item_dict[cal]:
                            # print(child_cal,col)
                            child_val += self.json_dict[child_cal][col]['value']

                        if round(parent_val) == round(child_val):
                            pass
                        else:
                            # print(parent_val,child_val,cal,col)
                            self.jsonerror_counter += 1
                            self.json_error[
                                self.jsonerror_counter] = f'Calculation of parent col {cal} doesnot match with child nodes {col}'


def load_financial_data_json(json_dict, tender_id, user_id):

    # deleting old tendering data from db
    TenderFinancialData.query.filter_by(ten_TenderId=tender_id).delete()
    db.session.commit()

    for ins in json_dict:
        local_item_name = (ins).lower().replace(' ', '')
        if ins not in ['item_name']:
            data = TenderFinancialData()
            data.ten_TenderId = tender_id
            data.ten_EquipmentCostItemId = json_dict['item_name'][local_item_name]['id']
            data.CreatedOn = datetime.now()
            data.CreatedBy=user_id
            for dat in json_dict[ins]:
                setattr(data, dat, json_dict[ins][dat]['value'])

            data.EstimatedCost = data.EquipmentEME + data.EquipmentInternal + data.EquipmentExternal + data.Fuel + data.MaterialsPipeline + data.MaterialsTeeth + data.MaterialsSteel + data.MaterialsRock + data.MaterialsImportedFill + data.MaterialsGeotextile \
                                 + data.MaterialsConcrete + data.MaterialsNavAid + data.MaterialsMarineFurniture + data.MaterialsOther + data.PersonalLabour + data.PersonalStaff + data.Misc + data.Subcontractor
            data.TotalCost = data.Contingencies + data.EstimatedCost
            data.MarkupsRisk = data.MarkupsRiskPercent * data.TotalCost
            data.MarkupsOverhead = data.MarkupsOverheadPercent * data.TotalCost
            data.MarkupsProfit = data.MarkupsProfitPercent * data.TotalCost
            data.TotalCost_exclTax = ((1 + data.MarkupsRiskPercent) / (
                        1 - data.MarkupsOverheadPercent - data.MarkupsProfitPercent)) * data.TotalCost
            data.TotalCost_inclTax = data.TotalCost_exclTax + data.TaxDirect + data.TaxIndirect
            db.session.add(data)
            db.session.commit()
    return {"Message": 'Added successfully'}


class LoadEquipmentDredging():
    def __init__(self, path):
        ### Set max font family value to 100 ###
        self.wb_obj = openpyxl.load_workbook(filename=path, data_only=True)
        self.act_budg_sheet = self.wb_obj['Act_Budget']
        self.dry_budg_sheet = self.wb_obj['DryEquipment']
        self.marine_budg_sheet = self.wb_obj['MarineEquipment']

        self.createdby = 'excel_import_pkr'
        self.createdon = datetime.now()
        self.json_dataform_dict = {'item_name':{}}
        self.json_dataform_error = {}
        self.jsonerror_dform_counter = 0

    def dredging_equip_data(self):
        equip_obj = EquipmentItem.query.all()
        equip_item_dict = {}
        equip_dict_name = {}
        for equip_ins in equip_obj:
            equip_item_dict[equip_ins.Item] = {'Id': equip_ins.Id , 'ParentId': equip_ins.ParentId,  'IsMarineEquipment': equip_ins.IsMarineEquipment}
            equip_dict_name[equip_ins.Id] = equip_ins.Item

        equip_row_ref = {}
        map_obj = EquipmentItemDrdgMap.query.all()
        for equip_map_ins in map_obj:
            equip_row_ref[equip_map_ins.DrdgEquipmentName] = equip_dict_name[equip_map_ins.ten_EquipmentItemId]

        equip_col_ref = {'D&I': 'EquipmentDI',
                    'Fixed M&R': 'EquipmentDIFixedMR',
                    'Variable M&R per OH': 'EquipmentDIVariableMR',
                    'Variable M&R per m3': 'EquipmentDIVariableMR',
                    'Insurance': 'EquipmentInsurance',
                    'Technical Overhead': 'MiscTechOH',
                    'Staff': 'PersonalStaff',
                    'Labour': 'PersonalLabour',
                    'External Equipment': 'External',
                    'Fuel': 'Fuel',
                    'Pipeline': 'Pipeline',
                    'Teeth': 'Teeth',
                    'Others': 'Others',
                    'Food': 'Food',
                    'Accomodation': 'Accomodation',
                    'Transport': 'Transport'}


        for category_row in self.act_budg_sheet['E2:E100000']:
            if category_row[0].value in ['Reclamation','Dredging']:
                if category_row[0].value not in self.json_dataform_dict:
                    self.json_dataform_dict[category_row[0].value] = {'eme':{},'marine':{}}
                
                ## adapted from old code
                equip_type = self.act_budg_sheet.cell(row= category_row[0].row, column = 4).value
                if equip_type in equip_row_ref:
                    sql_equip_type = equip_row_ref[equip_type]
                    if equip_item_dict[sql_equip_type]['IsMarineEquipment']:
                        marine_or_eme = 'marine'
                    else:
                        marine_or_eme = 'eme'

                    if sql_equip_type in equip_item_dict:
                        parent_id = equip_item_dict[sql_equip_type]['ParentId']

                        ## row wise dict to capture column values
                        if sql_equip_type not in self.json_dataform_dict:
                            ## Saving row wise id and status in the system
                            if sql_equip_type not in self.json_dataform_dict['item_name']:
                                self.json_dataform_dict['item_name'][sql_equip_type] = {'Id': equip_item_dict[sql_equip_type]['Id'], 'status':1}
                                self.json_dataform_dict[sql_equip_type] = {}

                            if parent_id:
                                if equip_dict_name[parent_id] not in self.json_dataform_dict['item_name']:
                                    self.json_dataform_dict['item_name'][equip_dict_name[parent_id]] = {'Id':parent_id, 'status':1}
                                    self.json_dataform_dict[equip_dict_name[parent_id]] = {}

                    else:
                        self.jsonerror_dform_counter +=1
                        self.json_dataform_error[self.jsonerror_dform_counter] =  f'Equipment type database - Data for {sql_equip_type} corresponding to excel name {category_row[0].value} not found in database. Check cell {category_row[0].coordinate}'     
                        
                    # equipment column data
                    cost_type = self.act_budg_sheet.cell(row= category_row[0].row, column = 6).value
                    TotalCost = self.act_budg_sheet.cell(row= category_row[0].row, column = 11).value
                    # check if the cost type from the excel present in the default excel mapped list
                    if cost_type in equip_col_ref:
                        sql_cost_type = equip_col_ref[cost_type]
                        # check the error for total cost
                        if isinstance(TotalCost,float) or isinstance(TotalCost,int):
                            # Create key based sql col name in the row wise dict if it is not created already
                            if sql_cost_type not in self.json_dataform_dict[category_row[0].value][marine_or_eme]:
                                    self.json_dataform_dict[category_row[0].value][marine_or_eme][sql_cost_type] = {'value': TotalCost, 'status':1}
                            else:
                                self.json_dataform_dict[category_row[0].value][marine_or_eme][sql_cost_type]['value']+=TotalCost


                            if sql_cost_type not in self.json_dataform_dict[sql_equip_type]:
                                    self.json_dataform_dict[sql_equip_type][sql_cost_type] = {'value': TotalCost, 'status':1}
                            else:
                                self.json_dataform_dict[sql_equip_type][sql_cost_type]['value']+=TotalCost

                            if parent_id:
                                if sql_cost_type not in self.json_dataform_dict[equip_dict_name[parent_id]]:
                                    self.json_dataform_dict[equip_dict_name[parent_id]][sql_cost_type] = {'value': TotalCost, 'status':1}
                                else:
                                    self.json_dataform_dict[equip_dict_name[parent_id]][sql_cost_type]['value']+=TotalCost

                        else:
                            self.jsonerror_dform_counter+=1
                            self.json_dataform_error[self.jsonerror_dform_counter] =  f'Total cost data - Cost Data is not numeric. Check cell {self.act_budg_sheet.cell(row= category_row[0].row, column = 11).coordinate}'
                            if sql_cost_type not in self.json_dataform_dict[category_row[0].value][marine_or_eme]:
                                self.json_dataform_dict[category_row[0].value][marine_or_eme][sql_cost_type] = {'value': 0, 'status':0}
                            else:
                                self.json_dataform_dict[category_row[0].value][marine_or_eme][sql_cost_type]['status'] = 0

                            if sql_cost_type not in self.json_dataform_dict[sql_equip_type]:
                                self.json_dataform_dict[sql_equip_type][sql_cost_type] = {'value': 0, 'status':0}
                            else:
                                self.json_dataform_dict[sql_equip_type][sql_cost_type]['status'] = 0
                                
                            if parent_id:
                                if sql_cost_type not in self.json_dataform_dict[equip_dict_name[parent_id]]:
                                    self.json_dataform_dict[equip_dict_name[parent_id]][sql_cost_type] = {'value': 0, 'status':0}
                                else:
                                    self.json_dataform_dict[equip_dict_name[parent_id]][sql_cost_type]['status']=0
                        

                    else:
                        self.jsonerror_dform_counter +=1
                        self.json_dataform_error[self.jsonerror_dform_counter ] =  f'Cost type mapping - No mapping data found for of excel cost type {cost_type}. Check cell {self.act_budg_sheet.cell(row= category_row[0].row, column = 6).coordinate}'

                
        ## calculate total internal, external and miscellaneous from individul json extracted from excel in the above code
        total_internal_dict = ['EquipmentDI','EquipmentDIFixedMR','EquipmentDIVariableMR','EquipmentInsurance']
        total_external_dict = ['External']
        misc_dict = ['MiscTechOH','Food','Accomodation','Transport']

        if self.jsonerror_dform_counter ==0:
            for k in self.json_dataform_dict:
                if k in ['Reclamation','Dredging']:
                    for k1 in self.json_dataform_dict[k]:
                        if 'total_internal' not in self.json_dataform_dict[k][k1]:
                            self.json_dataform_dict[k][k1]['total_internal'] = 0
                        for det in total_internal_dict:
                            if det in self.json_dataform_dict[k][k1]:
                                self.json_dataform_dict[k][k1]['total_internal'] += self.json_dataform_dict[k][k1][det]['value']

                        if 'total_external' not in self.json_dataform_dict[k][k1]:
                            self.json_dataform_dict[k][k1]['total_external'] = 0
                        for det in total_external_dict:
                            if det in self.json_dataform_dict[k][k1]:
                                self.json_dataform_dict[k][k1]['total_external'] += self.json_dataform_dict[k][k1][det]['value']

                        if 'total_misc' not in self.json_dataform_dict[k][k1]:
                            self.json_dataform_dict[k][k1]['total_misc'] = 0
                        for det in misc_dict:
                            if det in self.json_dataform_dict[k][k1]:
                                self.json_dataform_dict[k][k1]['total_misc'] += self.json_dataform_dict[k][k1][det]['value']
                                
                        if 'TotalCost' not in self.json_dataform_dict[k][k1]:
                            self.json_dataform_dict[k][k1]['TotalCost'] = 0
                        self.json_dataform_dict[k][k1]['TotalCost'] = self.json_dataform_dict[k][k1]['total_internal'] + self.json_dataform_dict[k][k1]['total_external']
                else:
                    if 'total_internal' not in self.json_dataform_dict[k]:
                        self.json_dataform_dict[k]['total_internal'] = 0
                    for det in total_internal_dict:
                        if det in self.json_dataform_dict[k]:
                            self.json_dataform_dict[k]['total_internal'] += self.json_dataform_dict[k][det]['value']

                    if 'total_external' not in self.json_dataform_dict[k]:
                        self.json_dataform_dict[k]['total_external'] = 0
                    for det in total_external_dict:
                        if det in self.json_dataform_dict[k]:
                            self.json_dataform_dict[k]['total_external'] += self.json_dataform_dict[k][det]['value']

                    if 'total_misc' not in self.json_dataform_dict[k]:
                        self.json_dataform_dict[k]['total_misc'] = 0
                    for det in misc_dict:
                        if det in self.json_dataform_dict[k]:
                            self.json_dataform_dict[k]['total_misc'] += self.json_dataform_dict[k][det]['value']


                    if 'TotalCost' not in self.json_dataform_dict[k]:
                        self.json_dataform_dict[k]['TotalCost'] = 0
                    
                    self.json_dataform_dict[k]['TotalCost'] = self.json_dataform_dict[k]['total_internal'] + self.json_dataform_dict[k]['total_external']

            ## No of equipments for eme and marine
            for col_row in self.dry_budg_sheet['E2:E100000']:
                if col_row[0].value in self.json_dataform_dict:
                    if 'NoOfEquipment' not in self.json_dataform_dict[col_row[0].value]:
                        self.json_dataform_dict[col_row[0].value]['NoOfEquipment'] = 0

                    include_units = self.dry_budg_sheet.cell(row= col_row[0].row, column = 14).value
                    if include_units == True:
                        self.json_dataform_dict[col_row[0].value]['NoOfEquipment']+=1

            for col_row in self.marine_budg_sheet['E2:E100000']:
                if col_row[0].value in self.json_dataform_dict:
                    if 'NoOfEquipment' not in self.json_dataform_dict[col_row[0].value]:
                        self.json_dataform_dict[col_row[0].value]['NoOfEquipment'] = 0

                    include_units = self.marine_budg_sheet.cell(row= col_row[0].row, column = 14).value
                    if include_units == True:
                        self.json_dataform_dict[col_row[0].value]['NoOfEquipment']+=1          


def load_dredging_eq_data_json(json_dataform_dict, tender_id, user_id):

    # deleting old tendering data from db
    TenderEquipmentCost.query.filter(
        TenderEquipmentCost.ten_Tender_Id == tender_id,
        TenderEquipmentCost.IsDredging == True
    ).delete()
    db.session.commit()

    tendereequip_table_col_list = [column.key for column in TenderEquipmentCost.__table__.columns]
    for formdict_ins in json_dataform_dict:
        if formdict_ins not in ['item_name', 'Reclamation', 'Dredging']:
            data = TenderEquipmentCost()
            data.ten_Tender_Id = tender_id
            data.IsDredging = True
            data.ten_EquipmentItemId = json_dataform_dict['item_name'][formdict_ins]['Id']
            data.CreatedOn = datetime.now()
            data.CreatedBy = user_id
            for dat in json_dataform_dict[formdict_ins]:
                if dat in tendereequip_table_col_list:
                    if dat in ['NoOfEquipment', 'TotalCost']:
                        setattr(data, dat, json_dataform_dict[formdict_ins][dat])
                    else:
                        setattr(data, dat, json_dataform_dict[formdict_ins][dat]['value'])

            if data.NoOfEquipment is None:
                data.NoOfEquipment = 0

            db.session.add(data)
            db.session.commit()

    return {"Message": 'Added successfully'}       


class LoadEquipmentCostEst():
    def __init__(self, path):
        ### Set max font family value to 100 ###
        self.wb_obj = openpyxl.load_workbook(filename=path, data_only=True)
        self.est_budg_sheet = self.wb_obj['Estimators']

        self.createdby = 'excel_import_pkr'
        self.createdon = datetime.now()
        self.json_eq_dict = {'item':{}}
        self.json_eq_error= {}
        self.json_eq_counter= 0

    def estimation_equip_data(self):
        equipitem_obj = EquipmentItem.query.all()
        equip_dict = {}
        for equipitem_ins in equipitem_obj:
            equip_dict[equipitem_ins.Item] = equipitem_ins.Id

        for est_ins in self.est_budg_sheet['A5:A10000']:
            equipment_item_name = est_ins[0].value
            if equipment_item_name is not None:
                if equipment_item_name in equip_dict:
                    if equipment_item_name not in self.json_eq_dict['item']:
                        self.json_eq_dict['item'][equipment_item_name] = {'Id': equip_dict[equipment_item_name], 'status': 1}


                    no_of_equipments_cell = self.est_budg_sheet.cell(row=est_ins[0].row,column=3)
                    total_cost_cell = self.est_budg_sheet.cell(row=est_ins[0].row,column=13)

                    if (total_cost_cell.value is not None and total_cost_cell.value >0) and (no_of_equipments_cell.value is not None  or no_of_equipments_cell.value >0) :
                        if isinstance(total_cost_cell.value, float) or isinstance(total_cost_cell.value, int):
                            if equipment_item_name not in self.json_eq_dict:
                                self.json_eq_dict[equipment_item_name] = {}
                            self.json_eq_dict[equipment_item_name]['TotalCost'] = {'value': total_cost_cell.value, 'status': 1}
                        else:
                            self.json_eq_counter+=1
                            self.json_eq_error[self.json_eq_counter] =  f'Data type of total cost is cannot float {total_cost_cell.coordinate}'

                        if isinstance(no_of_equipments_cell.value, float) or isinstance(no_of_equipments_cell.value, int):
                            if equipment_item_name not in self.json_eq_dict:
                                self.json_eq_dict[equipment_item_name] = {}
                            self.json_eq_dict[equipment_item_name]['NoOfEquipment'] = {'value': no_of_equipments_cell.value, 'status': 1}
                        else:
                            self.json_eq_counter+=1
                            self.json_eq_error[self.json_eq_counter] =  f'Data type of no of equipment cannot float {no_of_equipments_cell.coordinate}'
                    else:
                        
                        if total_cost_cell.value is None or total_cost_cell.value ==0:
                            if no_of_equipments_cell.value is not None:
                                if no_of_equipments_cell.value>0:
                                    self.json_eq_counter+=1
                                    self.json_eq_error[self.json_eq_counter] =  f'Number of Equipment should be None {no_of_equipments_cell.coordinate} when total cost is zero {total_cost_cell.coordinate}'

                        if no_of_equipments_cell.value is None or no_of_equipments_cell.value ==0:
                            if total_cost_cell.value is not None:
                                if total_cost_cell.value>0:
                                    self.json_eq_counter+=1
                                    self.json_eq_error[self.json_eq_counter] =  f'Number of Equipment should not be None  {no_of_equipments_cell.coordinate} when total cost has some value {total_cost_cell.coordinate}'
                else:
                    self.json_eq_counter+=1
                    self.json_eq_error[self.json_eq_counter] =  f'Number of Equipment should be None {no_of_equipments_cell.coordinate} when total cost is zero {total_cost_cell.coordinate}'


def load_estimator_eq_data_json(json_dataform_dict, tender_id, user_id):
    # deleting old tendering data from db
    TenderEquipmentCost.query.filter(
        TenderEquipmentCost.ten_Tender_Id == tender_id,
        or_(
            TenderEquipmentCost.IsDredging == None,
            TenderEquipmentCost.IsDredging == False,
        )
    ).delete()
    db.session.commit()

    tendereequip_table_col_list = [column.key for column in TenderEquipmentCost.__table__.columns]
    for formdict_ins in json_dataform_dict:
        if formdict_ins not in ['item']:
            data = TenderEquipmentCost()
            data.ten_Tender_Id = tender_id
            data.ten_EquipmentItemId = json_dataform_dict['item'][formdict_ins]['Id']
            data.CreatedOn = datetime.now()
            data.CreatedBy = user_id
            data.IsDredging = False
            for dat in json_dataform_dict[formdict_ins]:
                if dat in tendereequip_table_col_list:
                    setattr(data,dat,json_dataform_dict[formdict_ins][dat]['value'])
            db.session.add(data)
            db.session.commit()

    return {"Message": 'Added successfully'}


class LoadTenderPlanning():
    def __init__(self, path):
        ### Set max font family value to 100 ###
        self.workbook = openpyxl.load_workbook(path,data_only='True')

        # Access a specific worksheet by name
        self.worksheet = self.workbook['DataForm']

        self.createdby = 'excel_import_pkr'
        self.createdon = datetime.now()
        self.json_pl_dict = {'item':{}}
        self.json_pl_error= {}
        self.json_pl_counter= 0

    def planning_data(self, tender_id):
        act_obj = EquipmentCostItem.query.all()
        act_dict = {}
        act_ref_dict =  {}
        for act_ins in act_obj:
            if act_ins.ParentId == None:
                act_dict[act_ins.Id] = {'ParentId': act_ins.ParentId,'itemref':act_ins.ItemReference}

            if act_ins.Child != None:
                act_dict[act_ins.Id] = {'ParentId': act_ins.ParentId,'itemref':act_ins.ItemReference}

            if act_ins.SubChild != None:
                act_dict[act_ins.Id] = {'ParentId': act_ins.ParentId,'itemref':act_ins.ItemReference}

            act_ref_dict[act_ins.ItemReference] = act_ins.Id

        tender_financial_obj = TenderFinancialData.query.filter(TenderFinancialData.ten_TenderId==tender_id).all()
        fin_dict = {}
        for tender_financial_ins in tender_financial_obj:
            if tender_financial_ins.TotalCost_inclTax>0:
                fin_dict[tender_financial_ins.ten_EquipmentCostItemId] = tender_financial_ins.TotalCost_inclTax

        for i,j in enumerate(self.worksheet['B3:G67']):
            if j[4].value != None and j[5].value != None:
                # if j[0].value not in json_pl_dict:
                #     json_pl_dict[j[0].value] = {}
                if isinstance(j[4].value,datetime) and isinstance(j[4].value,datetime):
                    activity_code = (j[2].value).replace(' ','')
                    if activity_code not in act_ref_dict:
                        self.json_pl_counter+=1
                        self.json_pl_error[self.json_pl_counter] =  f'Defined activity code not present. Check cell {j[2].coordinate}'
                    else:
                        ## Current DB id
                        id = act_ref_dict[activity_code]
                        self.json_pl_dict['item'][activity_code]=id
                        if id in fin_dict:
                            if act_dict[id]['ParentId'] is not None:
                                parent_ref = act_dict[act_dict[id]['ParentId']]['itemref']
                                if parent_ref in self.json_pl_dict:
                                    min_date = self.json_pl_dict[parent_ref]['DateStart']
                                    max_date = self.json_pl_dict[parent_ref]['DateFinish']
                                    if min_date<=j[4].value and max_date>=j[5].value:
                                        self.json_pl_dict[activity_code] = {'DateStart':j[4].value,'DateFinish':j[5].value}
                                    else:
                                        self.json_pl_counter+=1
                                        self.json_pl_error[self.json_pl_counter] =  f'Parent data binding not enclosed. Check cell {j[2].coordinate}'
                                else:
                                    self.json_pl_counter+=1
                                    self.json_pl_error[self.json_pl_counter] =  f'Order Mismatch. Check cell {j[2].coordinate}'
                            else:            
                                self.json_pl_dict[activity_code] = {'DateStart':j[4].value,'DateFinish':j[5].value}
                        else:
                            self.json_pl_counter+=1
                            self.json_pl_error[self.json_pl_counter] =  f'There is no financial cost for the item {activity_code}. Check cell {j[2].coordinate}'
                else:
                    self.json_pl_counter+=1
                    self.json_pl_error[self.json_pl_counter] =  f'Data type is not integer. Check cell {j[2].coordinate}'


def load_planning_data_json(json_pl_dict, tender_id, user_id):
    # deleting old tendering data from db
    Planning.query.filter(
        Planning.ten_Tender_Id == tender_id,
    ).delete()
    db.session.commit()

    dates = []

    for json_pl_ins in json_pl_dict:
        if json_pl_ins not in ['item']:
            data = Planning()
            data.ten_Tender_Id = tender_id
            data.ten_EquipmentCostItemId = json_pl_dict['item'][json_pl_ins]
            data.CreatedOn = datetime.now()
            data.CreatedBy = user_id
            data.DateStart = json_pl_dict[json_pl_ins]['DateStart']
            data.DateFinish = json_pl_dict[json_pl_ins]['DateFinish']
            db.session.add(data)
            db.session.commit()

            dates.append(json_pl_dict[json_pl_ins]['DateStart'])
            dates.append(json_pl_dict[json_pl_ins]['DateFinish'])

    if len(dates) > 0:
        start = min(dates)
        end = max(dates)

        tender = Tenders.query.filter(Tenders.Id==tender_id).first()
        if tender:
            tender.AnticipatedStartDate = start
            tender.AnticipatedEndDate = end
            db.session.commit()


    return {"Message": 'Added successfully'}


def upload_risk_register_excel(path, tender_id, user_id):

    workbook = openpyxl.load_workbook(filename=path, data_only=True)
    wsheet = workbook['Register']

    levelOfRisk = LevelOfRisk.query.all()
    riskCategory = RiskCategory.query.all()

    riskCategory = {x.Category.lower(): x.Id for x in riskCategory}

    levelOfRisk_dist = {}
    for risk in levelOfRisk:
        if risk.Likelihood not in levelOfRisk_dist:
            levelOfRisk_dist[risk.Likelihood] = {}
        if risk.Consequence not in levelOfRisk_dist[risk.Likelihood]:
            levelOfRisk_dist[risk.Likelihood][risk.Consequence] = risk.Id


    data_to_insert = []
    json_error = []
    is_risk = False
    check = True

    current_row = 8
    for row in wsheet.iter_rows(min_row=8):
        if row[0].value !=None and row[1].value !=None:
            if row[1].value== 'Risks':
                is_risk = True
            elif row[1].value == 'Opportunities':
                is_risk = False
            else:
                check = False
                json_error.append('Error in selection of Risk/opportunity. Check cell B' + str(current_row))
                current_row+=1
                continue
            
            if wsheet.cell(current_row, 1).value == 'Variable':
                isfixed = 0
            elif wsheet.cell(current_row, 1).value == 'Fixed':
                isfixed = 1
            else:
                check = False
                json_error.append('Error in selection of Fixed/Variable. Check cell A' + str(current_row))
                current_row+=1
                continue

            category = wsheet.cell(current_row, 3).value.strip().replace(" ", "").lower()
            if category in riskCategory:
                risk_cat_id = riskCategory[category]
            else:
                check = False
                json_error.append('Risk Category not present in the system. Check cell C' + str(current_row))
                risk_cat_id = None
            
            if is_risk:
                
                if isinstance(wsheet.cell(current_row, 7).value,int):
                    if wsheet.cell(current_row, 7).value <5:
                        likehd = wsheet.cell(current_row, 7).value
                    else:
                        check = False
                        json_error.append('Likelihood cannot be more than 4. G' + str(current_row))
                        likehd = 0
                else:
                    check = False
                    json_error.append('Likelihood is not integer. G' + str(current_row))
                    likehd = 0

                if isinstance(wsheet.cell(current_row, 8).value,int):
                    if wsheet.cell(current_row, 8).value <5:
                        consq = wsheet.cell(current_row, 8).value
                    else:
                        check = False
                        json_error.append('Consequence cannot be more than 4. H' + str(current_row))
                        consq = 0
                else:
                    check = False
                    json_error.append('Consequence is not integer. H' + str(current_row))
                    consq = 0

                if likehd!=0 and consq!=0:
                    levelriskid = levelOfRisk_dist[likehd][consq]
                else:
                    levelriskid = None


                if isinstance(wsheet.cell(current_row, 11).value,int):
                    if wsheet.cell(current_row, 11).value <5:
                        reslikehd = wsheet.cell(current_row, 11).value
                    else:
                        check = False
                        json_error.append('Residual Likelihood cannot be more than 4. K' + str(current_row))
                        reslikehd = 0
                else:
                    check = False
                    json_error.append('Residual Likelihood is not integer. K' + str(current_row))
                    reslikehd = 0

                if isinstance(wsheet.cell(current_row, 12).value,int):
                    if wsheet.cell(current_row, 12).value <5:
                        resconsq = wsheet.cell(current_row, 12).value
                    else:
                        check = False
                        json_error.append('Residual Consequence cannot be more than 4. L' + str(current_row))
                        resconsq = 0
                else:
                    check = False
                    json_error.append('Residual Consequence is not integer. L' + str(current_row))
                    resconsq = 0

                if reslikehd!=0 and resconsq!=0:
                    reslevelriskid = levelOfRisk_dist[reslikehd][resconsq]
                else:
                    reslevelriskid = None
            else:
                levelriskid =None
                reslevelriskid =None


            if isinstance(wsheet.cell(current_row, 15).value,int) or isinstance(wsheet.cell(current_row, 15).value,float):
                riskvalue = wsheet.cell(current_row, 15).value
            else:
                check = False
                json_error.append('Risk value is not numeric. O' + str(current_row))
                riskvalue = None

            if isinstance(wsheet.cell(current_row, 16).value,int) or isinstance(wsheet.cell(current_row, 16).value,float):
                prob = wsheet.cell(current_row, 16).value
            else:
                check = False
                json_error.append('Probability is not numeric. P' + str(current_row))
                prob = None

            if isinstance(wsheet.cell(current_row, 17).value,int) or isinstance(wsheet.cell(current_row, 17).value,float):
                allowance = wsheet.cell(current_row, 17).value
            else:
                check = False
                json_error.append('Probability is not numeric. Q' + str(current_row))
                allowance = None

            data_to_insert.append(
                {
                    'ten_TenderId': tender_id,
                    'IsRisk': is_risk,
                    'IsFixed': isfixed,
                    'ten_RiskCategory':risk_cat_id ,
                    'RiskElement': wsheet.cell(current_row, 4).value,
                    'EventDetails': wsheet.cell(current_row, 5).value,
                    'ConsequenceDetails': wsheet.cell(current_row, 6).value,
                    'ten_LevelOfRiskId_initial': levelriskid,
                    'TreatmentMitigation': wsheet.cell(current_row, 10).value,
                    'ten_LevelOfRiskId_residual': reslevelriskid,
                    'ImplementationStatus': wsheet.cell(current_row, 14).value,
                    'RiskValue': riskvalue,
                    'Probability': prob * 100,
                    'RiskAllowance': allowance,
                    'CreatedBy': user_id,
                    'CreatedOn': datetime.now(),
                    'UpdatedBy': user_id,
                    'UpdatedOn': datetime.now(),
                }
            )
        current_row+=1

    return {
        'json_error': json_error,
        'data_to_insert': data_to_insert
    }


def in8_to_dataform_val(in8_data_path, new_dataform_path):
    wb_obj = openpyxl.load_workbook(filename=in8_data_path, data_only=True)
    res_sheet = wb_obj['Formatted Data']

    new_obj = openpyxl.load_workbook(filename=new_dataform_path, data_only=True)
    dataf_sheet = new_obj['DataForm']

    org_obj = OrganizationalCategory.query.all()
    org_dict = {}
    for org_ins in org_obj:
        org_dict[org_ins.OrganizationalCategoryItem] = org_ins.ExcelColumnName

    fin_data_ref_dict = {'EquipmentEME': 3,
                         'EquipmentInternal': 4,
                         'EquipmentExternal': 5,
                         'Fuel': 6,
                         'MaterialsPipeline': 7,
                         'MaterialsTeeth': 8,
                         'MaterialsSteel': 9,
                         'MaterialsRock': 10,
                         'MaterialsImportedFill': 11,
                         'MaterialsGeotextile': 12,
                         'MaterialsConcrete': 13,
                         'MaterialsNavAid': 14,
                         'MaterialsMarineFurniture': 15,
                         'MaterialsOther': 16,
                         'PersonalLabour': 18,
                         'PersonalStaff': 19,
                         'Misc': 20,
                         'Subcontractor': 21,
                         'Contingencies': 23,
                         'MarkupsRiskPercent': 25,
                         'MarkupsOverheadPercent': 26,
                         'MarkupsProfitPercent': 27,
                         'TaxDirect': 29,
                         'TaxIndirect': 30,
                         'Quantity': 34,
                         'Unit': 33,
                         'Remarks': 37}

    dataform_row_dict = {}
    for d in dataf_sheet['A10:A80']:
        if d[0].value is not None:
            dataform_row_dict[d[0].value] = d[0].row

    equipment_cost_obj = EquipmentCostItem.query.all()
    equipment_cost_dict = {}
    equip_dict = {}
    all_ids = []
    sub_child = []
    child = []
    for equip_ins in equipment_cost_obj:
        equipment_cost_dict[(equip_ins.ItemReference).replace(' ', '')] = equip_ins.Id
        equip_dict[equip_ins.Id] = {'name': equip_ins.ItemReference, 'parentid': equip_ins.ParentId}
        all_ids.append(equip_ins.Id)
        if equip_ins.Child is not None:
            child.append(equip_ins.Id)
        if equip_ins.SubChild is not None:
            sub_child.append(equip_ins.Id)

    for c in child:
        if equip_dict[c]['parentid'] in all_ids:
            all_ids.remove(equip_dict[c]['parentid'])

    for s in sub_child:
        if equip_dict[s]['parentid'] in all_ids:
            all_ids.remove(equip_dict[s]['parentid'])

    jsonerror_counter = 0
    json_error = {}
    dataf_load_json = {}
    for dat in res_sheet['A2:A100000']:
        if dat[0].value is not None:
            costcode = res_sheet.cell(row=dat[0].row, column=3).value
            org_cat = res_sheet.cell(row=dat[0].row, column=4).value
            col_name = None
            # Finding the column name through organisational category
            if org_cat not in org_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Organisation category {org_cat} doesnot exist in the database. Check cell D{dat[0].row}'
            else:
                if org_dict[org_cat] not in fin_data_ref_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'column name for {org_dict[org_cat]} does not exist in excel. Check cell D{dat[0].row}'
                else:
                    col_name = fin_data_ref_dict[org_dict[org_cat]]

            # checking the column name in databse
            row_name = None
            if costcode not in equipment_cost_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Cost code {costcode} doesnot exist in the database. Check cell C{dat[0].row}'
            else:
                if costcode not in dataform_row_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'Row name for {costcode} does not exist in excel. Check cell C{dat[0].row}'
                else:
                    row_name = dataform_row_dict[costcode]

            if col_name is not None and row_name is not None:
                # creating new row in the load json
                if row_name not in dataf_load_json:
                    dataf_load_json[row_name] = {}

                # create a new column under each row and add the cost
                if col_name not in dataf_load_json[row_name]:
                    cost = res_sheet.cell(row=dat[0].row, column=11).value
                    if cost is not None and (isinstance(cost, int) or isinstance(cost, float)):
                        dataf_load_json[row_name][col_name] = cost
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Total cost input is not in correct format. Check cell K{dat[0].row}'
                        dataf_load_json[row_name][col_name] = 0
                else:
                    cost = res_sheet.cell(row=dat[0].row, column=11).value
                    if cost is not None and (isinstance(cost, int) or isinstance(cost, float)):
                        dataf_load_json[row_name][col_name] += cost
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Total cost input is not in correct format. Check cell K{dat[0].row}'
                        dataf_load_json[row_name][col_name] += 0

    return json_error, dataf_load_json