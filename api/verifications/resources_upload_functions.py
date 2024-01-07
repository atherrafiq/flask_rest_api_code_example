import calendar
import openpyxl
from datetime import datetime, timedelta
from sqlalchemy import or_
from ..models.tenders_models import *
from openpyxl.utils.cell import coordinate_from_string


def load_manpower_resource_level(path, user_id):
    try:
        workbook = openpyxl.load_workbook(path, data_only='True')
        # Access a specific worksheet by name
        worksheet = workbook['Manpower_Resource']

        resource_lvl_obj = ResourceManpowerLevel.query.all()
        res_dict = {1:{},2:{},3:{},4:{}}
        for res_lvl_ins in resource_lvl_obj:
            if res_lvl_ins.LevelNumber not in res_dict:
                res_dict[res_lvl_ins.LevelNumber] = {}
            if res_lvl_ins.Name.lower().replace(" ","") not in res_dict[res_lvl_ins.LevelNumber]:
                res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ","") ] = res_lvl_ins.Id

        resource_dep_obj = ResourceDepartment.query.all()
        res_dept = {}
        for res_dept_ins in resource_dep_obj:
            if res_dept_ins.Name.lower().replace(" ","") not in res_dept:
                res_dept[res_dept_ins.Name.lower().replace(" ","")] = res_dept_ins.Id

        resource_code_obj = ResourceCode.query.all()
        res_code = {}
        for res_code_ins in resource_code_obj:
            if res_code_ins.Name.lower().replace(" ","") not in res_code:
                res_code[res_code_ins.Name.lower().replace(" ","")] = res_code_ins.Id

        # load level 1
        for d in worksheet['A2:A1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dict[1]:
                data = ResourceManpowerLevel(Name=d[0].value, LevelNumber=1, CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[1][(d[0].value).lower().replace(" ","") ] = 0

        # load level 2
        for d in worksheet['B2:B1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dict[2]:
                data = ResourceManpowerLevel(Name = d[0].value, LevelNumbe=2, CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[2][(d[0].value).lower().replace(" ","") ] = 0

        # load level 3
        for d in worksheet['C2:C1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dict[3]:
                data = ResourceManpowerLevel(Name=d[0].value, LevelNumber=3, CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[3][(d[0].value).lower().replace(" ","") ] = 0

        # load level 4
        for d in worksheet['D2:D1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dict[3]:
                data = ResourceManpowerLevel(Name = d[0].value, LevelNumber=4, CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[4][(d[0].value).lower().replace(" ","") ] = 0

        # Department
        for d in worksheet['E2:E1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dept:
                data = ResourceDepartment(Name = d[0].value, CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dept[(d[0].value).lower().replace(" ","") ] = 0

        # Resource code
        for d in worksheet['F2:F1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_code:
                data = ResourceCode(Name = d[0].value, Tag='Manpower', CreatedBy=user_id, CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_code[(d[0].value).lower().replace(" ","")] = 0

        return 200
    except:
        return 500


def load_manpower_resource(path, user_id):
    workbook = openpyxl.load_workbook(path, data_only='True')
    # Access a specific worksheet by name
    manpower_samp_sheet = workbook['Manpower_Sample']

    resource_lvl_obj = ResourceManpowerLevel.query.all()
    res_dict = {}
    for res_lvl_ins in resource_lvl_obj:
        if res_lvl_ins.LevelNumber not in res_dict:
            res_dict[res_lvl_ins.LevelNumber] = {}
        if res_lvl_ins.Name.lower().replace(" ", "") not in res_dict[res_lvl_ins.LevelNumber]:
            res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ", "")] = res_lvl_ins.Id

    res_man_obj = ResourceManpower.query.all()
    man_dict = {}
    for res_ins in res_man_obj:
        man_dict[res_ins.Name.lower().replace(" ", "")] = res_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    res_dept = {}
    for res_dept_ins in resource_dep_obj:
        if res_dept_ins.Name.lower().replace(" ", "") not in res_dept:
            res_dept[res_dept_ins.Name.lower().replace(" ", "")] = res_dept_ins.Id

    resource_code_obj = ResourceCode.query.all()
    res_code = {}
    for res_code_ins in resource_code_obj:
        if res_code_ins.Name.lower().replace(" ", "") not in res_code:
            res_code[res_code_ins.Name.lower().replace(" ", "")] = res_code_ins.Id

    error_dict = {}
    error_counter = 0
    # load level 1
    for d in manpower_samp_sheet['A2:A1000000']:
        if d[0].value is not None:

            manpower_name = d[0].value
            level1_val = manpower_samp_sheet.cell(row=d[0].row, column=2).value
            level2_val = manpower_samp_sheet.cell(row=d[0].row, column=3).value
            level3_val = manpower_samp_sheet.cell(row=d[0].row, column=4).value
            level4_val = manpower_samp_sheet.cell(row=d[0].row, column=5).value
            department_val = manpower_samp_sheet.cell(row=d[0].row, column=6).value
            resourcecode_val = manpower_samp_sheet.cell(row=d[0].row, column=7).value

            if manpower_name.lower().replace(" ", "") not in man_dict:
                manpower_str = str(manpower_name)

                if level1_val is not None:
                    if level1_val.lower().replace(" ", "") in res_dict[1]:
                        level1_id = res_dict[1][level1_val.lower().replace(" ", "")]
                    else:
                        level1_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 1 is not registered in database. check cell B{row_idx}'
                else:
                    level1_id = None

                if level2_val is not None:
                    if level2_val.lower().replace(" ", "") in res_dict[2]:
                        level2_id = res_dict[2][level2_val.lower().replace(" ", "")]
                    else:
                        level2_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 2 is not registered in database. check cell C{row_idx}'
                else:
                    level2_id = None

                if level3_val is not None:
                    if level3_val.lower().replace(" ", "") in res_dict[3]:
                        level3_id = res_dict[3][level3_val.lower().replace(" ", "")]
                    else:
                        level3_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 3 is not registered in database. check cell D{row_idx}'
                else:
                    level3_id = None

                if level4_val is not None:
                    if level4_val.lower().replace(" ", "") in res_dict[4]:
                        level4_id = res_dict[4][level4_val.lower().replace(" ", "")]
                    else:
                        level4_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 4 is not registered in database. check cell E{row_idx}'
                else:
                    level4_id = None

                if department_val is not None:
                    if department_val.lower().replace(" ", "") in res_dept:
                        dep_id = res_dept[department_val.lower().replace(" ", "")]
                    else:
                        dep_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[
                            error_counter] = f'Department name is not registered in database. check cell F{row_idx}'
                else:
                    dep_id = None

                if resourcecode_val is not None:
                    if resourcecode_val.lower().replace(" ", "") in res_code:
                        rescode_id = res_code[resourcecode_val.lower().replace(" ", "")]
                    else:
                        rescode_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[
                            error_counter] = f'Resource code is not registered in database. check cell G{row_idx}'
                else:
                    rescode_id = None

                if level1_id != 0 and level2_id != 0 and level3_id != 0 and level4_id != 0 and dep_id != 0 and rescode_id != 0:
                    data = ResourceManpower(
                        Name=manpower_str,
                        ten_ResourceManpowerLevelId_1=level1_id,
                        ten_ResourceManpowerLevelId_2=level2_id,
                        ten_ResourceManpowerLevelId_3=level3_id,
                        ten_ResourceManpowerLevelId_4=level4_id,
                        ten_ResourceDepartmentId=dep_id,
                        ten_ResourceCodeId=rescode_id,
                        CreatedBy=user_id,
                        CreatedOn=datetime.now()
                    )
                    db.session.add(data)
                    db.session.commit()
                    man_dict[manpower_name.lower().replace(" ", "")] = 0
            else:
                error_counter += 1
                row_idx = coordinate_from_string(d[0].coordinate)[1]
                error_dict[
                    error_counter] = f'Specified Name {manpower_name} already exist in the database. check cell F{row_idx}'
    return error_dict


def load_equipment_resource_level(path, user_id):
    try:
        workbook = openpyxl.load_workbook(path, data_only='True')
        # Access a specific worksheet by name
        worksheet = workbook['Equipment_Resource']

        resource_lvl_obj = ResourceEquipmentLevel.query.all()
        res_dict = {1: {}, 2: {}, 3: {}}
        for res_lvl_ins in resource_lvl_obj:
            if res_lvl_ins.LevelNumber not in res_dict:
                res_dict[res_lvl_ins.LevelNumber] = {}
            if res_lvl_ins.Name.lower().replace(" ", "") not in res_dict[res_lvl_ins.LevelNumber]:
                res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ", "")] = res_lvl_ins.Id

        resource_dep_obj = ResourceDepartment.query.all()
        res_dept = {}
        for res_dept_ins in resource_dep_obj:
            if res_dept_ins.Name.lower().replace(" ", "") not in res_dept:
                res_dept[res_dept_ins.Name.lower().replace(" ", "")] = res_dept_ins.Id

        resource_code_obj = ResourceCode.query.all()
        res_code = {}
        for res_code_ins in resource_code_obj:
            if res_code_ins.Name.lower().replace(" ", "") not in res_code:
                res_code[res_code_ins.Name.lower().replace(" ", "")] = res_code_ins.Id

        # load level 1
        for d in worksheet['A2:A1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ", "") not in res_dict[1]:
                data = ResourceEquipmentLevel(Name=d[0].value, LevelNumber=1,
                                              CreatedBy=user_id,
                                              CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[1][(d[0].value).lower().replace(" ", "")] = 0

        # load level 2
        for d in worksheet['B2:B1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ", "") not in res_dict[2]:
                data = ResourceEquipmentLevel(Name=d[0].value, LevelNumber=2,
                                              CreatedBy=user_id,
                                              CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[2][(d[0].value).lower().replace(" ", "")] = 0

        # load level 3
        for d in worksheet['C2:C1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ", "") not in res_dict[3]:
                data = ResourceEquipmentLevel(Name=d[0].value, LevelNumber=3,
                                              CreatedBy=user_id,
                                              CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[3][(d[0].value).lower().replace(" ", "")] = 0

        # Department
        for d in worksheet['D2:D1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ", "") not in res_dept:
                data = ResourceDepartment(Name=d[0].value, CreatedBy=user_id,
                                          CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dept[(d[0].value).lower().replace(" ", "")] = 0

        # Resource code
        for d in worksheet['E2:E1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ", "") not in res_code:
                data = ResourceCode(Name=d[0].value, Tag='Equipment', CreatedBy=user_id,
                                    CreatedOn=datetime.now())
                db.session.add(data)
                db.session.commit()
                res_code[(d[0].value).lower().replace(" ", "")] = 0

        return 200
    except:
        return 500


def load_equipment_resource(path, user_id):
    workbook = openpyxl.load_workbook(path, data_only='True')
    # Access a specific worksheet by name
    eq_samp_sheet = workbook['Equipment_Sample']

    resource_lvl_obj = ResourceEquipmentLevel.query.all()
    res_dict = {}
    for res_lvl_ins in resource_lvl_obj:
        if res_lvl_ins.LevelNumber not in res_dict:
            res_dict[res_lvl_ins.LevelNumber] = {}
        if res_lvl_ins.Name.lower().replace(" ", "") not in res_dict[res_lvl_ins.LevelNumber]:
            res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ", "")] = res_lvl_ins.Id

    res_eq_obj = ResourceEquipment.query.all()
    eq_dict = {}
    for res_ins in res_eq_obj:
        eq_dict[res_ins.Name.lower().replace(" ", "")] = res_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    res_dept = {}
    for res_dept_ins in resource_dep_obj:
        if res_dept_ins.Name.lower().replace(" ", "") not in res_dept:
            res_dept[res_dept_ins.Name.lower().replace(" ", "")] = res_dept_ins.Id

    resource_code_obj = ResourceCode.query.all()
    res_code = {}
    for res_code_ins in resource_code_obj:
        if res_code_ins.Name.lower().replace(" ", "") not in res_code:
            res_code[res_code_ins.Name.lower().replace(" ", "")] = res_code_ins.Id

    error_dict = {}
    error_counter = 0
    # load level 1
    for d in eq_samp_sheet['A2:A1000000']:
        if d[0].value is not None:
            eq_name = d[0].value
            level1_val = eq_samp_sheet.cell(row=d[0].row, column=2).value
            level2_val = eq_samp_sheet.cell(row=d[0].row, column=3).value
            level3_val = eq_samp_sheet.cell(row=d[0].row, column=4).value
            department_val = eq_samp_sheet.cell(row=d[0].row, column=5).value
            resourcecode_val = eq_samp_sheet.cell(row=d[0].row, column=6).value

            if eq_name.lower().replace(" ", "") not in eq_dict:
                eq_str = str(eq_name)

                if level1_val is not None:
                    if level1_val.lower().replace(" ", "") in res_dict[1]:
                        level1_id = res_dict[1][level1_val.lower().replace(" ", "")]
                    else:
                        level1_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 1 is not registered in database. check cell B{row_idx}'
                else:
                    level1_id = None

                if level2_val is not None:
                    if level2_val.lower().replace(" ", "") in res_dict[2]:
                        level2_id = res_dict[2][level2_val.lower().replace(" ", "")]
                    else:
                        level2_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 2 is not registered in database. check cell C{row_idx}'
                else:
                    level2_id = None

                if level3_val is not None:
                    if level3_val.lower().replace(" ", "") in res_dict[3]:
                        level3_id = res_dict[3][level3_val.lower().replace(" ", "")]
                    else:
                        level3_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 3 is not registered in database. check cell D{row_idx}'
                else:
                    level3_id = None

                if department_val is not None:
                    if department_val.lower().replace(" ", "") in res_dept:
                        dep_id = res_dept[department_val.lower().replace(" ", "")]
                    else:
                        dep_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[
                            error_counter] = f'Department name is not registered in database. check cell E{row_idx}'
                else:
                    dep_id = None

                if resourcecode_val is not None:
                    if resourcecode_val.lower().replace(" ", "") in res_code:
                        rescode_id = res_code[resourcecode_val.lower().replace(" ", "")]
                    else:
                        rescode_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[
                            error_counter] = f'Resource code is not registered in database. check cell F{row_idx}'
                else:
                    rescode_id = None

                if level1_id != 0 and level2_id != 0 and level3_id != 0 and dep_id != 0 and rescode_id != 0:
                    data = ResourceEquipment(Name=eq_str,
                                             ten_ResourceEquipmentLevelId_1=level1_id,
                                             ten_ResourceEquipmentLevelId_2=level2_id,
                                             ten_ResourceEquipmentLevelId_3=level3_id,
                                             ten_ResourceDepartmentId=dep_id,
                                             ten_ResourceCodeId=rescode_id,
                                             CreatedBy=user_id,
                                             CreatedOn=datetime.now())
                    db.session.add(data)
                    db.session.commit()
                    eq_dict[eq_name.lower().replace(" ", "")] = 0
            else:
                error_counter += 1
                row_idx = coordinate_from_string(d[0].coordinate)[1]
                error_dict[
                    error_counter] = f'Specified Name {eq_name} already exist in the database. check cell F{row_idx}'
    return error_dict


def load_material_resource_level(path, user_id):
    try:
        workbook = openpyxl.load_workbook(path,data_only='True')
        # Access a specific worksheet by name
        worksheet = workbook['Material_Resource']

        resource_lvl_obj = ResourceMaterialLevel.query.all()
        res_dict = {1:{}}
        for res_lvl_ins in resource_lvl_obj:
            if res_lvl_ins.LevelNumber not in res_dict:
                res_dict[res_lvl_ins.LevelNumber] = {}
            if res_lvl_ins.Name.lower().replace(" ","") not in res_dict[res_lvl_ins.LevelNumber]:
                res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ","") ] = res_lvl_ins.Id

        unit_obj = Unit.query.all()
        unit_dept = {}
        for unit_dept_ins in unit_obj:
            if unit_dept_ins.Name.lower().replace(" ","") not in unit_dept:
                unit_dept[unit_dept_ins.Name.lower().replace(" ","")] = unit_dept_ins.Id

        # load level 1
        for d in worksheet['A2:A1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in res_dict[1]:
                data = ResourceMaterialLevel(Name = d[0].value, LevelNumber =1, CreatedBy = user_id, CreatedOn = datetime.now())
                db.session.add(data)
                db.session.commit()
                res_dict[1][(d[0].value).lower().replace(" ","") ] = 0

        # Unit
        for d in worksheet['B2:B1000000']:
            if d[0].value is not None and (d[0].value).lower().replace(" ","") not in unit_dept:
                data = Unit(Name = d[0].value, CreatedBy = user_id, CreatedOn = datetime.now())
                db.session.add(data)
                db.session.commit()
                unit_dept[(d[0].value).lower().replace(" ","") ] = 0

        return 200
    except:
        return 500


def load_material_resource(path, user_id):
    workbook = openpyxl.load_workbook(path, data_only='True')
    # Access a specific worksheet by name
    material_samp_sheet = workbook['Material_Sample']

    resource_lvl_obj = ResourceMaterialLevel.query.all()
    res_dict = {}
    for res_lvl_ins in resource_lvl_obj:
        if res_lvl_ins.LevelNumber not in res_dict:
            res_dict[res_lvl_ins.LevelNumber] = {}
        if res_lvl_ins.Name.lower().replace(" ", "") not in res_dict[res_lvl_ins.LevelNumber]:
            res_dict[res_lvl_ins.LevelNumber][res_lvl_ins.Name.lower().replace(" ", "")] = res_lvl_ins.Id

    res_man_obj = ResourceMaterial.query.all()
    man_dict = {}
    for res_ins in res_man_obj:
        man_dict[res_ins.Name.lower().replace(" ", "")] = res_ins.Id

    unit_obj = Unit.query.all()
    unit_dict = {}
    for unit_ins in unit_obj:
        if unit_ins.Name.lower().replace(" ", "") not in unit_dict:
            unit_dict[unit_ins.Name.lower().replace(" ", "")] = unit_ins.Id

    error_dict = {}
    error_counter = 0
    # load level 1
    for d in material_samp_sheet['A2:A1000000']:
        if d[0].value is not None:

            material_name = d[0].value
            level1_val = material_samp_sheet.cell(row=d[0].row, column=2).value
            unit_val = material_samp_sheet.cell(row=d[0].row, column=3).value

            if material_name.lower().replace(" ", "") not in man_dict:
                material_str = str(material_name)

                if level1_val is not None:
                    if level1_val.lower().replace(" ", "") in res_dict[1]:
                        level1_id = res_dict[1][level1_val.lower().replace(" ", "")]
                    else:
                        level1_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[error_counter] = f'Level 1 is not registered in database. check cell B{row_idx}'
                else:
                    level1_id = None

                if unit_val is not None:
                    if unit_val.lower().replace(" ", "") in unit_dict:
                        unit_id = unit_dict[unit_val.lower().replace(" ", "")]
                    else:
                        unit_id = 0
                        error_counter += 1
                        row_idx = coordinate_from_string(d[0].coordinate)[1]
                        error_dict[
                            error_counter] = f'Department name is not registered in database. check cell F{row_idx}'
                else:
                    unit_id = None

                if level1_id != 0 and unit_id != 0:
                    data = ResourceMaterial(Name=material_str,
                                            ten_ResourceMaterialLevelId_1=level1_id,
                                            dbo_UnitId=unit_id,
                                            CreatedBy=user_id,
                                            CreatedOn=datetime.now())
                    db.session.add(data)
                    db.session.commit()
                    man_dict[material_name.lower().replace(" ", "")] = 0
            else:
                error_counter += 1
                row_idx = coordinate_from_string(d[0].coordinate)[1]
                error_dict[
                    error_counter] = f'Specified Name {material_name} already exist in the database. check cell F{row_idx}'
    return error_dict


def manpower_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    manpowersheet = workbook['Manpower R&D']

    # data from master table
    resource_manpower_obj = ResourceManpower.query.all()
    resource_man_dict = {}
    for resource_man_ins in resource_manpower_obj:
        resource_man_dict[resource_man_ins.Name.lower().replace(" ", "")] = resource_man_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    resource_dep_dict = {}
    for resource_dep_ins in resource_dep_obj:
        resource_dep_dict[resource_dep_ins.Psix_Reference] = resource_dep_ins.Name.lower().replace(" ", "")

    # getting min and max date from the excel
    date_list = []
    for dat in manpowersheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                date_list.append(dat.value.date())

    min_date = min(date_list)
    max_date = max(date_list)

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    date_range_dict = {}
    # saving date range with resource id for column entries
    for dat in manpowersheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                if dat.value.date() == min_date:
                    if dat.value.month == 12:
                        end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                    else:
                        end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                    date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                elif dat.value.date() == max_date:
                    start_date = datetime(dat.value.year, dat.value.month, 1)
                    date_range_dict[dat.column] = {'DateStart': start_date, 'DateEnd': dat.value.date()}
                else:
                    if dat.value.day == 1:
                        if dat.value.month == 12:
                            end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                        else:
                            end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                        date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} is not starting from first of the month'
            else:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} not present in datetime format'

    # saving the json dict histogram values
    for dat in manpowersheet['A3:A100000']:
        if dat[0].value is not None:
            dep_abb = manpowersheet.cell(row=dat[0].row, column=2).value
            if dep_abb not in resource_dep_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Department for {dat[0].value} in B{dat[0].row} not present in database'
            else:
                manpower_comb_name = dat[0].value.lower().replace(" ", "") + '-' + resource_dep_dict[dep_abb]

            if manpower_comb_name not in resource_man_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
            else:
                if resource_man_dict[manpower_comb_name] not in json_dict:
                    json_dict[resource_man_dict[manpower_comb_name]] = {}
                for date_col in date_range_dict:
                    cell_val = manpowersheet.cell(row=dat[0].row, column=date_col).value
                    if isinstance(cell_val, int) or isinstance(cell_val, float):
                        if cell_val > 0:
                            st_date = date_range_dict[date_col]['DateStart']
                            e_date = date_range_dict[date_col]['DateEnd']
                            if (st_date, e_date) not in json_dict[resource_man_dict[manpower_comb_name]]:
                                json_dict[resource_man_dict[manpower_comb_name]][(st_date, e_date)] = cell_val
                            else:
                                json_dict[resource_man_dict[manpower_comb_name]][(st_date, e_date)] += cell_val
                            # json_dict[resource_man_dict[dat[0].value.lower().replace(" ","")]].append([date_range_dict[date_col]['DateStart'],date_range_dict[date_col]['DateEnd'],cell_val])
                    elif cell_val is None:
                        pass
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Data format for {dat[0].value} in column {date_col} is not numeric'
            # else:
            #     jsonerror_counter+=1
            #     json_error[jsonerror_counter] =  f'Duplicate record found for {dat[0].value} in {dat[0].coordinate}'

    return json_error, json_dict


def manpower_p6_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    # Access a specific worksheet by name
    manpowersheet = workbook['Manpower-P6']

    # data from master table
    resource_manpower_obj = ResourceManpower.query.all()
    resource_man_dict = {}
    for resource_man_ins in resource_manpower_obj:
        resource_man_dict[resource_man_ins.Name.lower().replace(" ", "")] = resource_man_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    resource_dep_dict = {}
    for resource_dep_ins in resource_dep_obj:
        resource_dep_dict[resource_dep_ins.Psix_Reference] = resource_dep_ins.Name.lower().replace(" ", "")

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    for dat in manpowersheet['A2:A10000']:
        if dat[0].value is not None:
            # if dat[0].value == 'Operator':
            # print(dat[0].value)

            # Date list
            if isinstance(manpowersheet.cell(row=dat[0].row, column=3).value, datetime):
                start_date = manpowersheet.cell(row=dat[0].row, column=3).value
            else:
                start_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'Start date for {dat[0].value} in C{dat[0].row} is not datetime format'

            if isinstance(manpowersheet.cell(row=dat[0].row, column=4).value, datetime):
                end_date = manpowersheet.cell(row=dat[0].row, column=4).value
            else:
                end_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'End date for {dat[0].value} in D{dat[0].row}is not datetime format'

            if start_date is not None and end_date is not None:
                # print(start_date,type(start_date))
                date_list = []
                act_date = start_date
                while start_date <= end_date:
                    date_list.append(start_date)
                    start_date = start_date.replace(day=28) + timedelta(days=4)
                    start_date = start_date.replace(day=1)

                if act_date.month == date_list[-1].month:
                    date_list.append(end_date)
                elif end_date.month == date_list[-1].month:
                    date_list[-1] = end_date

                min_date = min(date_list)
                max_date = max(date_list)
                date_range_dict = []
                if len(date_list) == 2 and (date_list[0].month == date_list[1].month):
                    date_range_dict.append({'DateStart': date_list[0], 'DateEnd': date_list[1]})
                else:
                    for i, dat1 in enumerate(date_list):
                        if dat1 == min_date:
                            if dat1.month == 12:
                                end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                            else:
                                end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                            date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})

                        elif dat1 == max_date:
                            start_date1 = datetime(dat1.year, dat1.month, 1)
                            date_range_dict.append({'DateStart': start_date1, 'DateEnd': dat1})
                        else:
                            if dat1.day == 1:
                                if dat1.month == 12:
                                    end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                                else:
                                    end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                                date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})
                # else:
                #     jsonerror_counter+=1
                #     json_error[jsonerror_counter] =  f'Date split for min and max is not correct. Check development team'

                dep_abb = manpowersheet.cell(row=dat[0].row, column=2).value
                if dep_abb not in resource_dep_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'Department for {dat[0].value} in B{dat[0].row} not present in database'
                else:
                    manpower_comb_name = dat[0].value.lower().replace(" ", "") + '-' + resource_dep_dict[dep_abb]

                    if manpower_comb_name not in resource_man_dict:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
                    else:
                        if resource_man_dict[manpower_comb_name] not in json_dict:
                            json_dict[resource_man_dict[manpower_comb_name]] = {}

                        budgetunits = manpowersheet.cell(row=dat[0].row, column=5).value
                        if budgetunits is not None:
                            units = float(budgetunits.strip('/d'))
                        else:
                            units = 0
                            jsonerror_counter += 1
                            json_error[
                                jsonerror_counter] = f'Units of {dat[0].value} in E{dat[0].row} not in recognised format'

                        if units > 0:
                            for date_col in date_range_dict:
                                if date_col['DateStart'].year == date_col['DateEnd'].year and date_col[
                                    'DateStart'].month == date_col['DateEnd'].month:
                                    diff = date_col['DateEnd'] - date_col['DateStart']

                                    if diff.seconds > 0:
                                        days_in_range = diff.days + (diff.seconds / 3600 / 24)
                                    else:
                                        days_in_range = diff.days + 1

                                    month_diff = datetime(date_col['DateStart'].year, date_col['DateStart'].month,
                                                          calendar.monthrange(date_col['DateStart'].year,
                                                                              date_col['DateStart'].month)[
                                                              1]) - datetime(date_col['DateStart'].year,
                                                                             date_col['DateStart'].month, 1)
                                    days_in_month = month_diff.days + (month_diff.seconds / 3600 / 24) + 1

                                    # days_in_month = (date(date_col['DateStart'].year, date_col['DateStart'].month, 1) + relativedelta(months=1)) - date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                    # days_in_range = (date_col['DateEnd'] - date_col['DateStart']).days + 1
                                    ratio = days_in_range / days_in_month
                                    cell_val = ratio * units

                                    if cell_val > 0:
                                        st_date = date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                        e_date = date(date_col['DateStart'].year, date_col['DateStart'].month,
                                                      calendar.monthrange(date_col['DateStart'].year,
                                                                          date_col['DateStart'].month)[1])
                                        if (st_date, e_date) not in json_dict[resource_man_dict[manpower_comb_name]]:
                                            json_dict[resource_man_dict[manpower_comb_name]][
                                                (st_date, e_date)] = cell_val
                                        else:
                                            json_dict[resource_man_dict[manpower_comb_name]][
                                                (st_date, e_date)] += cell_val

                                else:
                                    jsonerror_counter += 1
                                    json_error[jsonerror_counter] = f'Date split is not correct. Check development team'

            else:
                continue

    return json_error, json_dict


def manpower_hist_json_to_db(json_dict,tender_id, user_id):

    ManpowerHistogram.query.filter(ManpowerHistogram.ten_TendersId == tender_id).delete()
    db.session.commit()
    for ins in json_dict:
        for d_list in json_dict[ins]:
            data = ManpowerHistogram()
            data.ten_TendersId = tender_id
            data.ten_ResourceManpowerId = ins
            data.DateStart = d_list[0]
            data.DateFinish = d_list[1]
            data.Value = json_dict[ins][d_list]
            data.CreatedBy = user_id
            data.CreatedOn = datetime.now()
            db.session.add(data)
            db.session.commit()
    return True


def equipment_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    equipmentsheet = workbook['Equipment R&D']

    # data from master table
    resource_eq_obj = ResourceEquipment.query.all()
    resource_eq_dict = {}
    for resource_eq_ins in resource_eq_obj:
        resource_eq_dict[resource_eq_ins.Name.lower().replace(" ", "")] = resource_eq_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    resource_dep_dict = {}
    for resource_dep_ins in resource_dep_obj:
        resource_dep_dict[resource_dep_ins.Psix_Reference] = resource_dep_ins.Name.lower().replace(" ", "")

    # getting min and max date from the excel
    date_list = []
    for dat in equipmentsheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                date_list.append(dat.value.date())

    min_date = min(date_list)
    max_date = max(date_list)

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    date_range_dict = {}
    # saving date range with resource id for column entries
    for dat in equipmentsheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                if dat.value.date() == min_date:
                    if dat.value.month == 12:
                        end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                    else:
                        end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                    date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                elif dat.value.date() == max_date:
                    start_date = datetime(dat.value.year, dat.value.month, 1)
                    date_range_dict[dat.column] = {'DateStart': start_date, 'DateEnd': dat.value.date()}
                else:
                    if dat.value.day == 1:
                        if dat.value.month == 12:
                            end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                        else:
                            end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                        date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} is not starting from first of the month'
            else:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} not present in datetime format'

    # saving the json dict histogram values
    for dat in equipmentsheet['A3:A100000']:
        if dat[0].value is not None:
            dep_abb = equipmentsheet.cell(row=dat[0].row, column=2).value
            if dep_abb not in resource_dep_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Department for {dat[0].value} in B{dat[0].row} not present in database'
            else:
                eq_comb_name = dat[0].value.lower().replace(" ", "") + '-' + resource_dep_dict[dep_abb]

                if eq_comb_name not in resource_eq_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
                else:
                    if resource_eq_dict[eq_comb_name] not in json_dict:
                        json_dict[resource_eq_dict[eq_comb_name]] = {}
                    for date_col in date_range_dict:
                        cell_val = equipmentsheet.cell(row=dat[0].row, column=date_col).value
                        if isinstance(cell_val, int) or isinstance(cell_val, float):
                            if cell_val > 0:
                                st_date = date_range_dict[date_col]['DateStart']
                                e_date = date_range_dict[date_col]['DateEnd']
                                if (st_date, e_date) not in json_dict[resource_eq_dict[eq_comb_name]]:
                                    json_dict[resource_eq_dict[eq_comb_name]][(st_date, e_date)] = cell_val
                                else:
                                    json_dict[resource_eq_dict[eq_comb_name]][(st_date, e_date)] += cell_val
                                # json_dict[resource_man_dict[dat[0].value.lower().replace(" ","")]].append([date_range_dict[date_col]['DateStart'],date_range_dict[date_col]['DateEnd'],cell_val])
                        elif cell_val is None:
                            pass
                        else:
                            jsonerror_counter += 1
                            json_error[
                                jsonerror_counter] = f'Data format for {dat[0].value} in column {date_col} is not numeric'

    return json_error, json_dict


def equipment_p6_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    # Access a specific worksheet by name
    equipmentsheet = workbook['Plant-P6']

    # data from master table
    resource_eq_obj = ResourceEquipment.query.all()
    resource_eq_dict = {}
    for resource_eq_ins in resource_eq_obj:
        resource_eq_dict[resource_eq_ins.Name.lower().replace(" ", "")] = resource_eq_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    resource_dep_dict = {}
    for resource_dep_ins in resource_dep_obj:
        resource_dep_dict[resource_dep_ins.Psix_Reference] = resource_dep_ins.Name.lower().replace(" ", "")

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    for dat in equipmentsheet['A2:A10000']:
        if dat[0].value is not None:
            # if dat[0].value == 'Operator':
            # print(dat[0].value)

            # Date list
            if isinstance(equipmentsheet.cell(row=dat[0].row, column=3).value, datetime):
                start_date = equipmentsheet.cell(row=dat[0].row, column=3).value
            else:
                start_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'Start date for {dat[0].value} in C{dat[0].row} is not datetime format'

            if isinstance(equipmentsheet.cell(row=dat[0].row, column=4).value, datetime):
                end_date = equipmentsheet.cell(row=dat[0].row, column=4).value
            else:
                end_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'End date for {dat[0].value} in D{dat[0].row}is not datetime format'

            if start_date is not None and end_date is not None:
                # print(start_date,type(start_date))
                date_list = []
                act_date = start_date
                while start_date <= end_date:
                    date_list.append(start_date)
                    start_date = start_date.replace(day=28) + timedelta(days=4)
                    start_date = start_date.replace(day=1)

                if act_date.month == date_list[-1].month:
                    date_list.append(end_date)
                elif end_date.month == date_list[-1].month:
                    date_list[-1] = end_date

                min_date = min(date_list)
                max_date = max(date_list)
                date_range_dict = []
                if len(date_list) == 2 and (date_list[0].month == date_list[1].month):
                    date_range_dict.append({'DateStart': date_list[0], 'DateEnd': date_list[1]})
                else:
                    for i, dat1 in enumerate(date_list):
                        if dat1 == min_date:
                            if dat1.month == 12:
                                end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                            else:
                                end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                            date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})

                        elif dat1 == max_date:
                            start_date1 = datetime(dat1.year, dat1.month, 1)
                            date_range_dict.append({'DateStart': start_date1, 'DateEnd': dat1})
                        else:
                            if dat1.day == 1:
                                if dat1.month == 12:
                                    end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                                else:
                                    end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                                date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})
                # else:
                #     jsonerror_counter+=1
                #     json_error[jsonerror_counter] =  f'Date split for min and max is not correct. Check development team'

                dep_abb = equipmentsheet.cell(row=dat[0].row, column=2).value
                if dep_abb not in resource_dep_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'Department for {dat[0].value} in B{dat[0].row} not present in database'
                else:
                    eq_comb_name = dat[0].value.lower().replace(" ", "") + '-' + resource_dep_dict[dep_abb]

                    if eq_comb_name not in resource_eq_dict:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
                    else:
                        if resource_eq_dict[eq_comb_name] not in json_dict:
                            json_dict[resource_eq_dict[eq_comb_name]] = {}

                        budgetunits = equipmentsheet.cell(row=dat[0].row, column=5).value
                        if budgetunits is not None:
                            units = float(budgetunits.strip('/d'))
                        else:
                            units = 0
                            jsonerror_counter += 1
                            json_error[
                                jsonerror_counter] = f'Units of {dat[0].value} in E{dat[0].row} not in recognised format'

                        if units > 0:
                            for date_col in date_range_dict:
                                if date_col['DateStart'].year == date_col['DateEnd'].year and date_col[
                                    'DateStart'].month == date_col['DateEnd'].month:
                                    diff = date_col['DateEnd'] - date_col['DateStart']

                                    if diff.seconds > 0:
                                        days_in_range = diff.days + (diff.seconds / 3600 / 24)
                                    else:
                                        days_in_range = diff.days + 1

                                    month_diff = datetime(date_col['DateStart'].year, date_col['DateStart'].month,
                                                          calendar.monthrange(date_col['DateStart'].year,
                                                                              date_col['DateStart'].month)[
                                                              1]) - datetime(date_col['DateStart'].year,
                                                                             date_col['DateStart'].month, 1)
                                    days_in_month = month_diff.days + (month_diff.seconds / 3600 / 24) + 1

                                    # days_in_month = (date(date_col['DateStart'].year, date_col['DateStart'].month, 1) + relativedelta(months=1)) - date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                    # days_in_range = (date_col['DateEnd'] - date_col['DateStart']).days + 1
                                    ratio = days_in_range / days_in_month
                                    cell_val = ratio * units

                                    if cell_val > 0:
                                        st_date = date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                        e_date = date(date_col['DateStart'].year, date_col['DateStart'].month,
                                                      calendar.monthrange(date_col['DateStart'].year,
                                                                          date_col['DateStart'].month)[1])
                                        if (st_date, e_date) not in json_dict[resource_eq_dict[eq_comb_name]]:
                                            json_dict[resource_eq_dict[eq_comb_name]][(st_date, e_date)] = cell_val
                                        else:
                                            json_dict[resource_eq_dict[eq_comb_name]][(st_date, e_date)] += cell_val

                                else:
                                    jsonerror_counter += 1
                                    json_error[jsonerror_counter] = f'Date split is not correct. Check development team'
            else:
                continue

    return json_error, json_dict


def equip_hist_json_to_db(json_dict, tender_id, user_id):

    EquipmentHistogram.query.filter(EquipmentHistogram.ten_TendersId == tender_id).delete()
    db.session.commit()
    for ins in json_dict:
        for d_list in json_dict[ins]:
            data = EquipmentHistogram()
            data.ten_TendersId = tender_id
            data.ten_ResourceEquipmentId = ins
            data.DateStart = d_list[0]
            data.DateFinish = d_list[1]
            data.Value = json_dict[ins][d_list]
            data.CreatedBy = user_id
            data.CreatedOn = datetime.now()
            db.session.add(data)
            db.session.commit()
    return True


def material_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    materialsheet = workbook['Material R&D']

    # data from master table
    resource_mat_obj = ResourceMaterial.query.all()
    resource_mat_dict = {}
    for resource_mat_ins in resource_mat_obj:
        resource_mat_dict[resource_mat_ins.Name.lower().replace(" ", "")] = resource_mat_ins.Id

    # getting min and max date from the excel
    date_list = []
    for dat in materialsheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                date_list.append(dat.value.date())

    min_date = min(date_list)
    max_date = max(date_list)

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    date_range_dict = {}
    # saving date range with resource id for column entries
    for dat in materialsheet['C2:ZZ2'][0]:
        if dat.value is not None:
            if isinstance(dat.value, datetime):
                if dat.value.date() == min_date:
                    if dat.value.month == 12:
                        end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                    else:
                        end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                    date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                elif dat.value.date() == max_date:
                    start_date = datetime(dat.value.year, dat.value.month, 1)
                    date_range_dict[dat.column] = {'DateStart': start_date, 'DateEnd': dat.value.date()}
                else:
                    if dat.value.day == 1:
                        if dat.value.month == 12:
                            end_date = datetime(dat.value.year + 1, 1, 1) + timedelta(days=-1)
                        else:
                            end_date = datetime(dat.value.year, dat.value.month + 1, 1) + timedelta(days=-1)
                        date_range_dict[dat.column] = {'DateStart': dat.value.date(), 'DateEnd': end_date.date()}
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} is not starting from first of the month'
            else:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Date for {dat.value} in {dat.coordinate} not present in datetime format'

    # saving the json dict histogram values
    for dat in materialsheet['A3:A100000']:
        if dat[0].value is not None:
            # dep_abb = materialsheet.cell(row=dat[0].row, column=2).value
            # if dep_abb not in resource_dep_dict:
            #     jsonerror_counter+=1
            #     json_error[jsonerror_counter] =  f'Department for {dat[0].value} in B{dat[0].row} not present in database'
            # else:
            mat_comb_name = dat[0].value.lower().replace(" ", "")

            if mat_comb_name not in resource_mat_dict:
                jsonerror_counter += 1
                json_error[
                    jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
            else:
                if resource_mat_dict[mat_comb_name] not in json_dict:
                    json_dict[resource_mat_dict[mat_comb_name]] = {}
                for date_col in date_range_dict:
                    cell_val = materialsheet.cell(row=dat[0].row, column=date_col).value
                    if isinstance(cell_val, int) or isinstance(cell_val, float):
                        if cell_val > 0:
                            st_date = date_range_dict[date_col]['DateStart']
                            e_date = date_range_dict[date_col]['DateEnd']
                            if (st_date, e_date) not in json_dict[resource_mat_dict[mat_comb_name]]:
                                json_dict[resource_mat_dict[mat_comb_name]][(st_date, e_date)] = cell_val
                            else:
                                json_dict[resource_mat_dict[mat_comb_name]][(st_date, e_date)] += cell_val
                            # json_dict[resource_man_dict[dat[0].value.lower().replace(" ","")]].append([date_range_dict[date_col]['DateStart'],date_range_dict[date_col]['DateEnd'],cell_val])
                    elif cell_val is None:
                        pass
                    else:
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Data format for {dat[0].value} in column {date_col} is not numeric'

    return json_error, json_dict


def material_p6_hist_json_validation(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only='True')
    # Access a specific worksheet by name
    materialsheet = workbook['Material-P6']

    # data from master table
    resource_mat_obj = ResourceMaterial.query.all()
    resource_mat_dict = {}
    for resource_mat_ins in resource_mat_obj:
        resource_mat_dict[resource_mat_ins.Name.lower().replace(" ", "")] = resource_mat_ins.Id

    resource_dep_obj = ResourceDepartment.query.all()
    resource_dep_dict = {}
    for resource_dep_ins in resource_dep_obj:
        resource_dep_dict[resource_dep_ins.Psix_Reference] = resource_dep_ins.Name.lower().replace(" ", "")

    json_dict = {}
    json_error = {}
    jsonerror_counter = 0
    for dat in materialsheet['A2:A10000']:
        if dat[0].value is not None:
            # if dat[0].value == 'Bollard':
            # print(dat[0].value)

            # Date list
            if isinstance(materialsheet.cell(row=dat[0].row, column=3).value, datetime):
                start_date = materialsheet.cell(row=dat[0].row, column=3).value
            else:
                start_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'Start date for {dat[0].value} in C{dat[0].row} is not datetime format'

            if isinstance(materialsheet.cell(row=dat[0].row, column=4).value, datetime):
                end_date = materialsheet.cell(row=dat[0].row, column=4).value
            else:
                end_date = None
                jsonerror_counter += 1
                json_error[jsonerror_counter] = f'End date for {dat[0].value} in D{dat[0].row}is not datetime format'

            if start_date is not None and end_date is not None:
                # print(start_date,type(start_date))
                date_list = []
                act_date = start_date
                while start_date <= end_date:
                    date_list.append(start_date)
                    start_date = start_date.replace(day=28) + timedelta(days=4)
                    start_date = start_date.replace(day=1)

                if act_date.month == date_list[-1].month:
                    date_list.append(end_date)
                elif end_date.month == date_list[-1].month:
                    date_list[-1] = end_date

                min_date = min(date_list)
                max_date = max(date_list)
                date_range_dict = []
                if len(date_list) == 2 and (date_list[0].month == date_list[1].month):
                    date_range_dict.append({'DateStart': date_list[0], 'DateEnd': date_list[1]})
                else:
                    for i, dat1 in enumerate(date_list):
                        if dat1 == min_date:
                            if dat1.month == 12:
                                end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                            else:
                                end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                            date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})

                        elif dat1 == max_date:
                            start_date1 = datetime(dat1.year, dat1.month, 1)
                            date_range_dict.append({'DateStart': start_date1, 'DateEnd': dat1})
                        else:
                            if dat1.day == 1:
                                if dat1.month == 12:
                                    end_date1 = datetime(dat1.year + 1, 1, 1) + timedelta(days=-1)
                                else:
                                    end_date1 = datetime(dat1.year, dat1.month + 1, 1, 23, 59) + timedelta(days=-1)
                                date_range_dict.append({'DateStart': dat1, 'DateEnd': end_date1})
                # else:
                #     jsonerror_counter+=1
                #     json_error[jsonerror_counter] =  f'Date split for min and max is not correct. Check development team'

                # dep_abb = materialsheet.cell(row=dat[0].row, column=2).value
                # if dep_abb not in resource_dep_dict:
                #     jsonerror_counter+=1
                #     json_error[jsonerror_counter] =  f'Department for {dat[0].value} in B{dat[0].row} not present in database'
                # else:
                mat_comb_name = dat[0].value.lower().replace(" ", "")

                if mat_comb_name not in resource_mat_dict:
                    jsonerror_counter += 1
                    json_error[
                        jsonerror_counter] = f'Data for {dat[0].value} in {dat[0].coordinate} not present in database'
                else:
                    if resource_mat_dict[mat_comb_name] not in json_dict:
                        json_dict[resource_mat_dict[mat_comb_name]] = {}

                    budgetunits = materialsheet.cell(row=dat[0].row, column=5).value
                    if budgetunits is not None:
                        units = float(budgetunits.strip('/d'))
                    else:
                        units = 0
                        jsonerror_counter += 1
                        json_error[
                            jsonerror_counter] = f'Units of {dat[0].value} in E{dat[0].row} not in recognised format'

                    if units > 0:
                        for date_col in date_range_dict:
                            if date_col['DateStart'].year == date_col['DateEnd'].year and date_col['DateStart'].month == \
                                    date_col['DateEnd'].month:
                                diff = date_col['DateEnd'] - date_col['DateStart']

                                if diff.seconds > 0:
                                    days_in_range = diff.days + (diff.seconds / 3600 / 24)
                                else:
                                    days_in_range = diff.days + 1

                                month_diff = datetime(date_col['DateStart'].year, date_col['DateStart'].month,
                                                      calendar.monthrange(date_col['DateStart'].year,
                                                                          date_col['DateStart'].month)[1]) - datetime(
                                    date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                days_in_month = month_diff.days + (month_diff.seconds / 3600 / 24) + 1

                                # days_in_month = (date(date_col['DateStart'].year, date_col['DateStart'].month, 1) + relativedelta(months=1)) - date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                # days_in_range = (date_col['DateEnd'] - date_col['DateStart']).days + 1
                                # print(days_in_range,units)
                                ratio = days_in_range / days_in_month
                                cell_val = days_in_range * units

                                if cell_val > 0:
                                    st_date = date(date_col['DateStart'].year, date_col['DateStart'].month, 1)
                                    e_date = date(date_col['DateStart'].year, date_col['DateStart'].month,
                                                  calendar.monthrange(date_col['DateStart'].year,
                                                                      date_col['DateStart'].month)[1])
                                    if (st_date, e_date) not in json_dict[resource_mat_dict[mat_comb_name]]:
                                        json_dict[resource_mat_dict[mat_comb_name]][(st_date, e_date)] = cell_val
                                    else:
                                        json_dict[resource_mat_dict[mat_comb_name]][(st_date, e_date)] += cell_val

                            else:
                                jsonerror_counter += 1
                                json_error[jsonerror_counter] = f'Date split is not correct. Check development team'


            else:
                continue

    return json_error, json_dict


def material_hist_json_to_db(json_dict,tender_id, user_id):
    try:
        MaterialHistogram.query.filter(MaterialHistogram.ten_TendersId == tender_id).delete()
        db.session.commit()
        for ins in json_dict:
            for d_list in json_dict[ins]:
                data = MaterialHistogram()
                data.ten_TendersId = tender_id
                data.ten_ResourceMaterialId = ins
                data.DateStart = d_list[0]
                data.DateFinish = d_list[1]
                data.Value = json_dict[ins][d_list]
                data.CreatedBy = user_id
                data.CreatedOn=datetime.now()
                db.session.add(data)
                db.session.commit()
        return True
    except:
        return False